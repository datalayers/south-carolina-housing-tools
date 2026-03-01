"""
extract_fields.py  –  SC TEB Application Field Extractor
Companion to extract_coords.py

Extracts structured data from two pages of each SC TEB application PDF:
  • Page 6  – Unit Details & Proposed Development Income
  • Page 9  – Development Costs

Outputs an Excel workbook with:
  • Sheet "Page6_Summary"   – one row per project, summary totals
  • Sheet "Page6_UnitMix"   – one row per unit-type row per project
  • Sheet "Page6_OtherIncome" – one row per other-income row per project
  • Sheet "Page9_DevCosts"  – one row per cost line item per project

Usage:
    python extract_fields.py [--root ROOT] [--output-dir DIR] [--log-dir DIR]
"""

import os
import re
import sys
import argparse
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _clean(val: str) -> str:
    return val.strip() if isinstance(val, str) else ""


def _num(val: str) -> str:
    """Strip commas/dollar-signs so numeric strings survive as-is (no conversion)."""
    return val.replace(",", "").replace("$", "").strip() if isinstance(val, str) else val


def _find_pdfs_recursive(start_dir: str) -> list[str]:
    pdfs = []
    for root, _, files in os.walk(start_dir):
        for fn in files:
            if fn.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, fn))
    return pdfs


def _find_tab1_dir(project_root: str) -> str | None:
    """
    Fuzzy-match the TAB 1 / Application subfolder, mirroring extract_dev_costs_to_excel logic.
    Returns the path of the best candidate, or None if not found.
    """
    candidates = []
    for dirpath, dirnames, _ in os.walk(project_root):
        for dname in dirnames:
            name = dname.lower()
            full = os.path.join(dirpath, dname)
            if "tab" in name and re.search(r"\b1\b", name) and ("app" in name or "application" in name):
                candidates.append(full)
                continue
            if name.startswith("tab 1") or name.startswith("tab1") or "tab 01" in name:
                candidates.append(full)
                continue
            if re.search(r"^\s*1\s*[-_ ]", name) and ("app" in name or "application" in name):
                candidates.append(full)
    if not candidates:
        return None
    # Prefer shallowest (fewest path components)
    candidates.sort(key=lambda p: len(p.split(os.sep)))
    return candidates[0]


def _score_pdf(pdf_path: str) -> int:
    """Score a PDF for likelihood of being the main TEB application."""
    score = 0
    name = os.path.basename(pdf_path).lower().strip()
    bad_words = [
        "waiver", "map", "railroad", "site", "zoning", "utility",
        "environmental", "plans", "specifications", "appraisal",
        "market study", "syndication", "architect", "engineer",
        "certification", "site control", "checklist", "entity",
        "agreement", "opinion", "certifications", "survey", "title",
        "easement", "phase i", "phase 1", "soil", "geotech",
    ]
    for w in bad_words:
        if w in name:
            score -= 12
    if "executed" in name:
        score -= 25
    if "application page" in name:
        score -= 15
    if "application" in name:
        score += 12
    elif "app" in name:
        score += 6
    if "teb" in name:
        score += 10
    if "teb" in name and ("app" in name or "application" in name):
        score += 20
    if name in ("2025 teb application.pdf", "2026 teb application.pdf"):
        score += 30
    try:
        size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        score += min(int(size_mb), 20)
        if size_mb < 1.0:
            score -= 10
    except Exception:
        pass
    # Content check: strongly prefer PDFs that contain the Page 6 anchor text
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:15]:
                txt = page.extract_text() or ""
                if re.search(r"Unit Details and Proposed Development Income", txt, re.I):
                    score += 60
                    break
    except Exception:
        pass
    return score


def pick_best_pdf(project_path: str) -> str | None:
    """
    Find the TAB 1 folder (fuzzy), then score all PDFs inside it.
    Falls back to searching the entire project folder if no TAB 1 found.
    """
    tab1_dir = _find_tab1_dir(project_path)
    pdfs = _find_pdfs_recursive(tab1_dir) if tab1_dir else []
    if not pdfs:
        pdfs = _find_pdfs_recursive(project_path)
    if not pdfs:
        return None
    scored = sorted(
        [((_score_pdf(p), os.path.getsize(p) if os.path.exists(p) else 0), p) for p in pdfs],
        reverse=True,
    )
    return scored[0][1]


# ---------------------------------------------------------------------------
# PAGE 6 PARSER
# ---------------------------------------------------------------------------

# Regex patterns
_SET_ASIDE_20_RE = re.compile(r"X\s+At least 20%", re.I)
_SET_ASIDE_40_RE = re.compile(r"X\s+At least 40%", re.I)
_INCOME_AVG_RE  = re.compile(r"X\s+Income averaging", re.I)
_LIHTC_DATE_RE  = re.compile(r"LIHTC:\s*([\d/]+)", re.I)
_HOME_DATE_RE   = re.compile(r"HOME:\s*([\d/]+)", re.I)
_DEV_NAME_RE    = re.compile(r"^(.+?)\s+(\d{1,2}/\d{1,2}/\d{4})\s*$", re.M)

# Summary totals anchors
_SUMM = {
    "total_bedrooms":      re.compile(r"Total\s+bedrooms\s*=\s*([\d,]+)", re.I),
    "total_residential_sqft": re.compile(r"Total\s+Residential\s+Sqft\s*=\s*([\d,]+)", re.I),
    "total_annual_income": re.compile(r"Total\s+Annual\s+Income\s*=\s*([\d,]+)", re.I),
    "total_li_units":      re.compile(r"Total\s+LI\s+Units\s*=\s*([\d,]+)", re.I),
    "total_mr_units":      re.compile(r"Total\s+MR\s+Units\s*=\s*([\d,]+)", re.I),
    "li_unit_pct":         re.compile(r"LI\s+Unit\s+Percentage\s*=\s*([\d.%]+)", re.I),
    "total_li_sqft":       re.compile(r"Total\s+LI\s+Sqft\s*=\s*([\d,]+)", re.I),
    "total_mr_sqft":       re.compile(r"Total\s+MR\s+Sqft\s*=\s*([\d,]+)", re.I),
    "li_sqft_pct":         re.compile(r"LI\s+Sqft\s+Percentage\s*=\s*([\d.%]+)", re.I),
    "total_common_sqft":   re.compile(r"Total\s+Common\s+Sqft:\s*([\d,]+)", re.I),
    "total_non_heated_sqft": re.compile(r"Total\s+Non-Heated\s+Sqft:\s*([\d,]+)", re.I),
    "total_development_sqft": re.compile(r"Total\s+Development\s+Sqft\s*=\s*([\d,]+)", re.I),
    "other_income_total":  re.compile(r"Totals:\s*([\d,]+\.\d{2})", re.I),
}


def _detect_set_aside(full_text: str) -> str:
    """Return the elected set-aside option."""
    # The 'X' marker appears just before the chosen option in the extracted text
    lines = full_text.splitlines()
    for i, line in enumerate(lines):
        if re.search(r"\bX\b", line):
            ctx = " ".join(lines[max(0, i-1):i+3])
            if "20%" in ctx and "50%" in ctx:
                return "20/50"
            if "40%" in ctx and "60%" in ctx:
                return "40/60"
            if "Income averaging" in ctx or "income averaging" in ctx:
                return "Income Averaging"
    # Fallback: check for X near option text
    if _SET_ASIDE_20_RE.search(full_text):
        return "20/50"
    if _SET_ASIDE_40_RE.search(full_text):
        return "40/60"
    if _INCOME_AVG_RE.search(full_text):
        return "Income Averaging"
    return ""


def _detect_col_bands(words: list, header_top: float) -> list[tuple]:
    """
    Dynamically build column band definitions by finding known header words.
    This handles layout variations across applicants (e.g. presence/absence of
    'Unit Utility Type' column) without any hardcoded x-coordinates.

    Returns list of (x0_min, x0_max, field_name) tuples sorted by x0_min.
    """
    # Map each header keyword to the field it anchors
    # We use the FIRST occurrence of each keyword within the header band
    ANCHOR_MAP = {
        "Beds":       "beds",
        "Baths":      "baths",
        "Footage":    "sqft",           # "Square Footage"
        "Proposed":   "proposed_rent",  # "Proposed Monthly Rent*"
        "Allowance":  "util_allowance", # "Utility Allowance"
        "Gross":      "gross_rent",     # "Gross Rent"
        "Maximum":    "max_allowable_rent",  # "Maximum Allowable Rent"
        "AMGI":       "amgi_pct",
        "Assistance": "assistance_type",
    }

    # Collect anchor x0 values from words within the header band (±30pt of header_top)
    # Exclude pure-numeric words (those are data values that leaked into the band)
    anchors: dict[str, float] = {}
    for w in words:
        if abs(w["top"] - header_top) < 30 and w["text"] in ANCHOR_MAP:
            if re.match(r"^[\d,\.]+$", w["text"]):
                continue  # skip numeric leakage
            field = ANCHOR_MAP[w["text"]]
            if field not in anchors:  # first occurrence only
                anchors[field] = w["x0"]

    # Find the '#' of Units column
    for w in words:
        if abs(w["top"] - header_top) < 30 and w["text"] == "#":
            anchors["num_units"] = w["x0"]
            break

    # Find Type column — first 'Type' word in header band (leftmost)
    for w in sorted(words, key=lambda x: x["x0"]):
        if abs(w["top"] - header_top) < 30 and w["text"] == "Type":
            if not re.match(r"^[\d,\.]+$", w["text"]):
                anchors["type"] = w["x0"]
                break

    # Find Unit Utility Type column — the 'Utility' left of 'Allowance'
    allow_x = anchors.get("util_allowance", 999)
    util_xs = [w["x0"] for w in words
               if abs(w["top"] - header_top) < 30
               and w["text"] == "Utility"
               and w["x0"] < allow_x - 50]
    if util_xs:
        anchors["unit_description"] = min(util_xs)

    # Sort anchors by x0, then build bands: each column extends from its x0
    # to the next column's x0 (minus a small gap)
    # row_num is always the leftmost column with no header word — hardcode ~28
    named_cols = sorted(anchors.items(), key=lambda x: x[1])
    all_cols = [("row_num", 28.0)] + [(name, x) for name, x in named_cols]

    bands = []
    for i, (name, x_start) in enumerate(all_cols):
        x_end = all_cols[i + 1][1] - 1 if i + 1 < len(all_cols) else 620
        # row_num gets a generous left margin; all others start tight at their anchor
        left_margin = 6 if name == "row_num" else 1
        bands.append((x_start - left_margin, x_end, name))

    return bands


def _parse_unit_table(page) -> list[dict]:
    """
    Extract the unit-type rows from Page 6.
    Dynamically detects column positions from header word x0 coordinates,
    making the parser robust to layout variations across applicants.
    """
    words = page.extract_words()

    # Find vertical range: header at 'Beds', bottom at 'Detail of Other Income'
    header_top = None
    table_bottom = None

    for w in words:
        if w["text"] == "Beds" and header_top is None:
            header_top = w["top"]
        if w["text"] == "Detail" and header_top is not None:
            idx = words.index(w)
            nearby = [words[j]["text"] for j in range(idx, min(idx+5, len(words)))]
            if "Other" in nearby:
                table_bottom = w["top"]
                break

    if header_top is None:
        return []

    # Build dynamic column bands from header word positions
    col_bands = _detect_col_bands(words, header_top)

    def assign_col(x0: float) -> str:
        for (lo, hi, name) in col_bands:
            if lo <= x0 < hi:
                return name
        return "unknown"

    # Collect data words: below header, above Other Income section
    HEADER_WORDS = {"Beds", "Baths", "Footage", "Allowance", "Rent*",
                    "Proposed", "Maximum", "#", "of", "Square", "Utility",
                    "Monthly", "Gross", "Rent", "Allowable", "%", "AMGI",
                    "Assistance", "Type", "Units", "Unit", "and", "Income",
                    "Detail", "Other"}
    top_limit = table_bottom if table_bottom else 9999
    # Use header_top + 18 to skip the multi-line header rows cleanly
    # but catch the first data row which may sit only ~20pt below beds_top
    data_words = [w for w in words
                  if w["top"] > header_top + 18
                  and w["top"] < top_limit
                  and w["text"] not in HEADER_WORDS]

    # Group words into rows by proximity: words within 2pt vertically belong together.
    # Sort by top, then merge each word into the nearest existing group if within 2pt,
    # otherwise start a new group. This handles both sub-pixel offsets (row_num vs data)
    # and proper row separation (~10pt between rows).
    sorted_data = sorted(data_words, key=lambda w: w["top"])
    rows_by_top: dict[float, list] = {}
    for w in sorted_data:
        matched = None
        for key in rows_by_top:
            if abs(w["top"] - key) <= 2.0:
                matched = key
                break
        if matched is None:
            matched = w["top"]
        rows_by_top.setdefault(matched, []).append(w)

    unit_rows = []
    for top_key in sorted(rows_by_top):
        row_words = sorted(rows_by_top[top_key], key=lambda w: w["x0"])
        row_dict: dict[str, list] = {}
        for w in row_words:
            col = assign_col(w["x0"])
            row_dict.setdefault(col, []).append(w["text"])

        row_num_val = " ".join(row_dict.get("row_num", []))
        if not re.match(r"^\d{1,2}$", row_num_val):
            continue
        rn = int(row_num_val)
        if rn < 1 or rn > 20:
            continue

        type_raw   = " ".join(row_dict.get("type", []))
        unit_val   = " ".join(row_dict.get("unit_description", []))
        # type should be just "LI" or "MR" — anything after that is the unit description
        type_parts = type_raw.split()
        if type_parts and type_parts[0] in ("LI", "MR"):
            type_val = type_parts[0]
            if not unit_val and len(type_parts) > 1:
                unit_val = " ".join(type_parts[1:])
        else:
            type_val = type_raw
        num_units  = " ".join(row_dict.get("num_units", []))
        beds       = " ".join(row_dict.get("beds", []))
        baths      = " ".join(row_dict.get("baths", []))
        sqft       = " ".join(row_dict.get("sqft", []))
        prop_rent  = " ".join(row_dict.get("proposed_rent", []))
        util_allow = " ".join(row_dict.get("util_allowance", []))
        gross_rent = " ".join(row_dict.get("gross_rent", []))
        max_rent   = " ".join(row_dict.get("max_allowable_rent", []))
        amgi_raw   = " ".join(row_dict.get("amgi_pct", []))
        asst_raw   = " ".join(row_dict.get("assistance_type", []))
        # amgi should be numeric; if it contains text (e.g. "60 Section"), split it
        amgi_parts = amgi_raw.split()
        amgi_overflow = []
        while amgi_parts and not re.match(r"^\d+$", amgi_parts[-1]):
            amgi_overflow.insert(0, amgi_parts.pop())
        amgi = " ".join(amgi_parts)
        asst = (" ".join(amgi_overflow) + " " + asst_raw).strip()

        # A valid unit row must have a recognised type (LI/MR) AND a unit count.
        # This prevents empty rows 2-20 from leaking through when a stray word
        # gets proximity-clustered into them.
        if type_val not in ("LI", "MR") or not num_units:
            continue

        unit_rows.append({
            "row_num":               rn,
            "type":                  type_val,
            "unit_description":      unit_val,
            "num_units":             _num(num_units),
            "beds":                  beds,
            "baths":                 baths,
            "sqft":                  _num(sqft),
            "proposed_monthly_rent": _num(prop_rent),
            "utility_allowance":     _num(util_allow),
            "gross_rent":            _num(gross_rent),
            "max_allowable_rent":    _num(max_rent),
            "amgi_pct":              amgi,
            "assistance_type":       asst,
        })

    return unit_rows


def _parse_other_income(page) -> list[dict]:
    """
    Extract Detail of Other Income rows (rows 1-7 beneath the 'Detail of Other Income' header).
    """
    words = page.extract_words()

    # Find anchor: 'Detail' word near 'Other' 'Income'
    anchor_top = None
    for i, w in enumerate(words):
        if w["text"] == "Detail":
            # Confirm 'Other' and 'Income' follow within same line
            nearby = [words[j]["text"] for j in range(i, min(i+5, len(words)))]
            if "Other" in nearby and "Income" in nearby:
                anchor_top = w["top"]
                break

    if anchor_top is None:
        return []

    # Data rows start ~30pt below anchor, span ~100pt
    row_top_min = anchor_top + 28
    row_top_max = anchor_top + 120

    data_words = [w for w in words if row_top_min < w["top"] < row_top_max]

    # Column bands for other income table:
    # row_num ~32-50, income_type ~50-200, num_units ~175-240,
    # annual_amount ~240-305, pct_of_units ~305-355,
    # monthly_per_unit ~355-400, annual_per_unit ~400-460
    OI_BANDS = [
        (28,   52,  "row_num"),
        (52,   178, "income_type"),
        (178,  242, "num_units"),
        (242,  308, "annual_amount"),
        (306,  356, "pct_of_units"),
        (356,  400, "monthly_per_unit"),
        (398,  460, "annual_per_unit"),
    ]

    def assign_oi_col(x0):
        for (lo, hi, name) in OI_BANDS:
            if lo <= x0 < hi:
                return name
        return "unknown"

    rows_by_top: dict[float, list] = {}
    for w in data_words:
        key = round(w["top"] / 3) * 3
        rows_by_top.setdefault(key, []).append(w)

    oi_rows = []
    for top_key in sorted(rows_by_top):
        row_words = sorted(rows_by_top[top_key], key=lambda w: w["x0"])
        row_dict: dict[str, list] = {}
        for w in row_words:
            col = assign_oi_col(w["x0"])
            row_dict.setdefault(col, []).append(w["text"])

        row_num_val = " ".join(row_dict.get("row_num", []))
        if not re.match(r"^\d$", row_num_val):
            continue
        rn = int(row_num_val)
        if rn < 1 or rn > 7:
            continue

        income_type   = " ".join(row_dict.get("income_type", []))
        num_units_val = " ".join(row_dict.get("num_units", []))
        annual_amt    = _num(" ".join(row_dict.get("annual_amount", [])))
        pct           = " ".join(row_dict.get("pct_of_units", []))
        monthly_pu    = " ".join(row_dict.get("monthly_per_unit", []))
        annual_pu     = " ".join(row_dict.get("annual_per_unit", []))

        # Skip blank rows
        if not any([income_type, annual_amt]):
            continue

        oi_rows.append({
            "row_num":          rn,
            "income_type":      income_type,
            "num_units":        num_units_val,
            "annual_amount":    annual_amt,
            "pct_of_units":     pct,
            "monthly_per_unit": monthly_pu,
            "annual_per_unit":  annual_pu,
        })

    return oi_rows


def parse_page6(pdf_path: str, project: str) -> dict:
    """
    Returns dict with keys:
      'summary'      → dict of scalar fields
      'unit_rows'    → list of dicts (unit mix)
      'other_income' → list of dicts
      'status'       → 'OK' or error string
    """
    result = {"summary": {}, "unit_rows": [], "other_income": [], "status": "not_found"}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                if not re.search(r"Unit\s+Details\s+and\s+Proposed\s+Development\s+Income", txt, re.I):
                    continue

                # ---- Summary scalars ----
                summary = {"project_folder": project, "pdf_path": pdf_path}

                # Development name & date from header
                m = re.search(r"^(.+?)\s+(\d{1,2}/\d{1,2}/\d{4})\s*$", txt, re.M)
                if m:
                    summary["development_name"] = _clean(m.group(1))
                    summary["application_date"]  = m.group(2)
                else:
                    summary["development_name"] = ""
                    summary["application_date"]  = ""

                summary["set_aside_election"] = _detect_set_aside(txt)

                m = _LIHTC_DATE_RE.search(txt)
                summary["lihtc_effective_date"] = m.group(1) if m else ""
                m = _HOME_DATE_RE.search(txt)
                summary["home_effective_date"] = m.group(1) if m else ""

                for field, rx in _SUMM.items():
                    m = rx.search(txt)
                    summary[field] = _num(m.group(1)) if m else ""

                result["summary"] = summary

                # ---- Unit mix rows ----
                result["unit_rows"] = _parse_unit_table(page)

                # ---- Other income rows ----
                result["other_income"] = _parse_other_income(page)

                result["status"] = "OK"
                return result

    except Exception as e:
        result["status"] = f"ERROR: {e}"
    return result


# ---------------------------------------------------------------------------
# PAGE 9 PARSER  (Development Costs)
# ---------------------------------------------------------------------------

# Each cost line item has a line number (1-80 range) and label.
# We capture both the "4% Basis (30%) - Acquisition" column and
# "4%/9% New/Rehab" column, plus the Summary of Const Cost Addm. column.

_COST_LINE_RE = re.compile(
    r"^\s*(\d{1,2})\s+(.+?)\s+([\d,]+(?:\.\d+)?|-)\s*([\d,]+(?:\.\d+)?|-)?\s*([\d,]+(?:\.\d+)?|-)?",
    re.M
)

# Named line items we specifically care about (label anchor → field name)
COST_ITEMS = {
    # Acquisition
    1:  "land",
    2:  "existing_structures",
    3:  "other_acquisition",
    # Site Work
    4:  "on_site_improvements",
    5:  "off_site_improvements",
    6:  "demolition",
    7:  "site_improvements",
    # Rehab / New Const
    8:  "new_construction",
    9:  "rehabilitation",
    10: "accessory_structures",
    11: "other_hard_construction",
    12: "furniture_fixtures_equipment",
    13: "contractor_contingency",
    14: "general_requirements",
    15: "contractor_profit",
    16: "contractor_overhead",
    # Professional Fees
    17: "architect_design",
    18: "architect_supervision",
    19: "engineering_fees",
    20: "survey",
    21: "real_estate_attorney",
    22: "tax_attorney",
    23: "accountant",
    24: "green_certification",
    25: "other_professional",
    # Construction Financing
    26: "const_loan_origination",
    27: "const_loan_interest",
    28: "const_loan_legal",
    29: "const_loan_credit_report",
    30: "const_loan_title_recording",
    31: "inspection_fees",
    32: "other_const_financing",
    # Construction Interim
    33: "construction_insurance",
    34: "performance_bond",
    35: "construction_taxes",
    36: "tap_impact_fees",
    37: "permitting_fees",
    38: "other_const_interim",
    # Permanent Financing
    39: "perm_loan_origination",
    40: "bond_premium",
    41: "credit_enhancement",
    42: "perm_loan_title_recording",
    43: "counsels_fee",
    44: "lenders_counsel",
    46: "credit_report",
    47: "mortgage_broker",
    48: "perm_loan_closing",
    49: "underwriter_discount",
    50: "perm_attorney_legal",
    51: "letter_of_credit",
    # Soft Costs
    52: "feasibility_study",
    53: "environmental_study",
    45: "appraisal_fees",
    54: "market_study",
    55: "sch_application_fee",
    56: "sch_market_study",
    57: "sch_plan_spec_review",
    58: "sch_tc_reservation",
    59: "sch_bond_issuance",
    60: "compliance_fees",
    61: "cost_certification",
    62: "tenant_relocation",
    63: "soil_testing",
    64: "physical_needs_assessment",
    65: "rent_up_expenses",
    # (64 and 65 are reused in the template — we capture both occurrences)
    # Syndication
    66: "organizational_expenses",
    67: "tax_opinion",
    68: "bridge_loan_fees",
    69: "syndication_fees",
    70: "other_syndication",
    # Developer Fees
    71: "developer_overhead",
    72: "developer_fee",
    73: "project_consultant_fee",
    74: "other_developer",
    # Reserves
    75: "operating_reserves",
    76: "other_reserves",
}

SECTION_ANCHORS = {
    "Acquisition":                 "acquisition",
    "Site Work":                   "site_work",
    "Rehabilitation and New Construction": "rehab_new_const",
    "Professional Fees":           "professional_fees",
    "Construction Financing":      "construction_financing",
    "Construction Interim Costs":  "construction_interim",
    "Permanent Financing":         "permanent_financing",
    "Soft Costs":                  "soft_costs",
    "Syndication Costs":           "syndication",
    "Developer Fees":              "developer_fees",
    "Project Reserves":            "reserves",
}

SECTION_TOTAL_RE = re.compile(
    r"^\s*([\d,]+(?:\.\d+)?)\s*-?\s*([\d,]+(?:\.\d+)?)?\s*$", re.M
)


def _parse_cost_line(line: str):
    """
    Parse a cost line.  Several formats exist:
      '1  Land  1,102,000'                       (label then one value)
      '4  On-Site Improvements  -  1,045,200  1,045,200'   (three values)
      '39 Permanent Loan Origination Fee 265,490'           (one value, perm financing)
      '9  Rehabilitation  8,340,920  -  8,340,920'
    Handles both multi-space (Palomino) and single-space (Highland Square) layouts.
    """
    # Strip leading whitespace then require: digits, whitespace, label text,
    # whitespace, then 1-3 value tokens (numbers or "-")
    NUM_OR_DASH = r'([\d,]+(?:\.\d+)?|-)'
    m = re.match(
        r'^\s*(\d{1,2})\s+(.+?)\s{1,}'   # line_num + label
        + NUM_OR_DASH                      # col1 (required)
        + r'(?:\s+' + NUM_OR_DASH + r')?'  # col2 (optional)
        + r'(?:\s+' + NUM_OR_DASH + r')?'  # col3 (optional)
        + r'\s*$',
        line
    )
    if not m:
        return None

    def clean_val(v):
        if v is None or v == '-':
            return ''
        return v.replace(',', '')

    # Label cleanup: strip trailing dashes and whitespace
    label = re.sub(r'\s*-\s*$', '', m.group(2)).strip()
    if not label:
        return None

    return (
        int(m.group(1)),
        label,
        clean_val(m.group(3)),
        clean_val(m.group(4)),
        clean_val(m.group(5)),
    )


def parse_page9(pdf_path: str, project: str) -> dict:
    """
    Returns dict with keys:
      'cost_rows' → list of dicts, one per parsed cost line
      'totals'    → dict with total_dev_cost, total_eligible_basis, total_ineligible
      'status'    → 'OK' or error string
    """
    result = {"cost_rows": [], "totals": {}, "status": "not_found"}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                if not re.search(r"Development\s+Costs", txt, re.I):
                    continue
                # Must also have numbered cost lines
                if not re.search(r"^\s*1\s+Land", txt, re.M | re.I):
                    continue

                cost_rows = []
                lines = txt.splitlines()
                current_section = ""

                for line in lines:
                    # Detect section headers
                    for anchor, sec_name in SECTION_ANCHORS.items():
                        if anchor.lower() in line.lower() and len(line.strip()) < 60:
                            current_section = sec_name
                            break

                    parsed = _parse_cost_line(line)
                    if parsed:
                        line_num, label, col1, col2, col3 = parsed
                        cost_rows.append({
                            "project_folder": project,
                            "pdf_path":       pdf_path,
                            "section":        current_section,
                            "line_num":       line_num,
                            "line_name":      COST_ITEMS.get(line_num, label),
                            "label_raw":      label,
                            "total_col":      col1,
                            "basis_4pct_col": col2,
                            "basis_9pct_col": col3,
                        })

                # Totals
                totals = {"project_folder": project, "pdf_path": pdf_path}
                m = re.search(r"TOTAL DEVELOPMENT COST\s+([\d,]+(?:\.\d+)?)", txt, re.I)
                totals["total_development_cost"] = _num(m.group(1)) if m else ""
                m = re.search(r"TOTAL ELIGIBLE BASIS\s+([\d,]+(?:\.\d+)?)", txt, re.I)
                totals["total_eligible_basis"] = _num(m.group(1)) if m else ""
                m = re.search(r"TOTAL INELIGIBLE COSTS\s+([\d,]+(?:\.\d+)?)", txt, re.I)
                totals["total_ineligible_costs"] = _num(m.group(1)) if m else ""

                # Column totals row
                m = re.search(r"COLUMN TOTALS\s+([\d,]+(?:\.\d+)?)\s*-?\s*([\d,]+(?:\.\d+)?)?", txt, re.I)
                if m:
                    totals["column_total_1"] = _num(m.group(1))
                    totals["column_total_2"] = _num(m.group(2)) if m.group(2) else ""

                result["cost_rows"] = cost_rows
                result["totals"]    = totals
                result["status"]    = "OK"
                return result

    except Exception as e:
        result["status"] = f"ERROR: {e}"
    return result


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", start_color="1F497D")   # dark blue
SUBHEAD_FILL  = PatternFill("solid", start_color="4F81BD")   # medium blue
ALT_ROW_FILL  = PatternFill("solid", start_color="DCE6F1")   # light blue
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHEAD_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
BODY_FONT     = Font(name="Arial", size=9)
BOLD_FONT     = Font(name="Arial", bold=True, size=9)

thin_side   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _write_header_row(ws, row: int, headers: list[str], fill=HEADER_FILL, font=HEADER_FONT):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data_row(ws, row: int, values: list, alt: bool = False):
    fill = ALT_ROW_FILL if alt else WHITE_FILL
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)


def build_excel(
    p6_summary_rows: list[dict],
    p6_unit_rows: list[dict],
    p6_other_rows: list[dict],
    output_path: str,
):
    wb = Workbook()

    # ---- Sheet 1: Page6_Summary ----
    ws1 = wb.active
    ws1.title = "Page6_Summary"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 40

    p6_summary_cols = [
        "project_folder", "development_name", "application_date", "set_aside_election",
        "lihtc_effective_date", "home_effective_date",
        "total_bedrooms", "total_residential_sqft", "total_annual_income",
        "total_li_units", "total_mr_units", "li_unit_pct",
        "total_li_sqft", "total_mr_sqft", "li_sqft_pct",
        "total_common_sqft", "total_non_heated_sqft", "total_development_sqft",
        "other_income_total", "status", "pdf_path",
    ]
    p6_summary_labels = [
        "Project Folder", "Development Name", "Application Date", "Set-Aside Election",
        "LIHTC Effective Date", "HOME Effective Date",
        "Total Bedrooms", "Total Residential Sqft", "Total Annual Income ($)",
        "Total LI Units", "Total MR Units", "LI Unit %",
        "Total LI Sqft", "Total MR Sqft", "LI Sqft %",
        "Total Common Sqft", "Total Non-Heated Sqft", "Total Development Sqft",
        "Other Income Total ($)", "Status", "PDF Path",
    ]
    _write_header_row(ws1, 1, p6_summary_labels)
    for i, row in enumerate(p6_summary_rows):
        _write_data_row(ws1, i + 2, [row.get(c, "") for c in p6_summary_cols], alt=(i % 2 == 1))
    _auto_width(ws1)

    # ---- Sheet 2: Page6_UnitMix ----
    ws2 = wb.create_sheet("Page6_UnitMix")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 40

    um_cols = [
        "project_folder", "row_num", "type", "unit_description",
        "num_units", "beds", "baths", "sqft",
        "proposed_monthly_rent", "utility_allowance", "gross_rent",
        "max_allowable_rent", "amgi_pct", "assistance_type",
    ]
    um_labels = [
        "Project Folder", "Row #", "Type", "Unit Description",
        "# Units", "Beds", "Baths", "Sqft",
        "Proposed Monthly Rent ($)", "Utility Allowance ($)", "Gross Rent ($)",
        "Max Allowable Rent ($)", "% AMGI", "Assistance Type",
    ]
    _write_header_row(ws2, 1, um_labels)
    for i, row in enumerate(p6_unit_rows):
        _write_data_row(ws2, i + 2, [row.get(c, "") for c in um_cols], alt=(i % 2 == 1))
    _auto_width(ws2)

    # ---- Sheet 3: Page6_OtherIncome ----
    ws3 = wb.create_sheet("Page6_OtherIncome")
    ws3.freeze_panes = "A2"
    ws3.row_dimensions[1].height = 40

    oi_cols = [
        "project_folder", "row_num", "income_type", "num_units",
        "annual_amount", "pct_of_units", "monthly_per_unit", "annual_per_unit",
    ]
    oi_labels = [
        "Project Folder", "Row #", "Income Type", "# Units",
        "Annual Amount ($)", "% of Units", "Monthly $ / Unit", "Annual $ / Unit",
    ]
    _write_header_row(ws3, 1, oi_labels)
    for i, row in enumerate(p6_other_rows):
        _write_data_row(ws3, i + 2, [row.get(c, "") for c in oi_cols], alt=(i % 2 == 1))
    _auto_width(ws3)

    wb.save(output_path)
    print(f"Excel written to: {output_path}")


# ---------------------------------------------------------------------------
# CLI + MAIN
# ---------------------------------------------------------------------------

def parse_args():
    default_root    = r"C:\Users\tanne\Downloads\SC_TEB\2025\extracted"
    script_dir      = os.path.abspath(os.path.dirname(__file__))
    default_out_dir = os.path.join(script_dir, "output")
    default_log_dir = os.path.join(script_dir, "logs")

    p = argparse.ArgumentParser(
        description="Extract Page 6 & Page 9 fields from SC TEB application PDFs."
    )
    p.add_argument("--root",       default=default_root,    help="Root folder containing project subfolders.")
    p.add_argument("--output-dir", default=default_out_dir, help="Output directory for Excel workbook.")
    p.add_argument("--log-dir",    default=default_log_dir, help="Directory for run logs.")
    return p.parse_args()


def main():
    args = parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    os.makedirs(args.log_dir,    exist_ok=True)

    run_ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_xlsx = os.path.join(args.output_dir, f"SC_TEB_Fields_EXTRACTED_{run_ts}.xlsx")
    latest_xlsx = os.path.join(args.output_dir, "SC_TEB_Fields_EXTRACTED_latest.xlsx")
    log_path    = os.path.join(args.log_dir,    f"fields_run_{run_ts}.log")

    log_file        = open(log_path, "w", encoding="utf-8")
    original_stdout = sys.stdout

    class Tee:
        def __init__(self, *s): self.streams = s
        def write(self, d):
            for s in self.streams: s.write(d); s.flush()
        def flush(self):
            for s in self.streams: s.flush()

    sys.stdout = Tee(original_stdout, log_file)

    try:
        print(f"Root:       {args.root}")
        print(f"Output dir: {args.output_dir}")
        print()

        if not os.path.isdir(args.root):
            raise FileNotFoundError(f'Root folder not found: "{args.root}"')

        p6_summary_rows = []
        p6_unit_rows    = []
        p6_other_rows   = []

        for project in sorted(os.listdir(args.root)):
            project_path = os.path.join(args.root, project)
            if not os.path.isdir(project_path):
                continue

            pdf_path = pick_best_pdf(project_path)
            if not pdf_path:
                print(f"[SKIP] {project} – no PDF found")
                p6_summary_rows.append({"project_folder": project, "status": "No PDF found"})
                p6_summary_rows.append({"project_folder": project, "status": "No PDF found"})
                continue

            # --- Page 6 ---
            p6 = parse_page6(pdf_path, project)
            if p6["status"] == "OK":
                p6_summary_rows.append(p6["summary"])
                for ur in p6["unit_rows"]:
                    p6_unit_rows.append({"project_folder": project, **ur})
                for oi in p6["other_income"]:
                    p6_other_rows.append({"project_folder": project, **oi})
                print(f"[P6 OK]   {project} – {len(p6['unit_rows'])} unit rows, "
                      f"{len(p6['other_income'])} other income rows")
            else:
                p6_summary_rows.append({"project_folder": project,
                                         "pdf_path": pdf_path,
                                         "status": p6["status"]})
                print(f"[P6 MISS] {project} – {p6['status']}")

        build_excel(p6_summary_rows, p6_unit_rows, p6_other_rows, output_xlsx)

        import shutil
        shutil.copy2(output_xlsx, latest_xlsx)

        print(f"\nProjects processed: {len(p6_summary_rows)}")
        print(f"Page 6 summary rows: {len(p6_summary_rows)}")
        print(f"Unit mix rows:       {len(p6_unit_rows)}")
        print(f"Other income rows:   {len(p6_other_rows)}")
        print(f"\nExcel:   {output_xlsx}")
        print(f"Latest:  {latest_xlsx}")
        print(f"Log:     {log_path}")

    finally:
        sys.stdout = original_stdout
        log_file.close()


if __name__ == "__main__":
    main()
