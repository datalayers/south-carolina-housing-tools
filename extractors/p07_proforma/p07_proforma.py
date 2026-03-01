"""
p07_proforma.py — SC TEB Application Page 7 Extractor
======================================================
Extracts the Proforma Income Statement from Page 7 of each SC TEB application PDF.

Output: single Excel workbook with one sheet — Proforma (one row per project)

Usage:
    python p07_proforma.py [--root ROOT] [--output-dir DIR] [--log-dir DIR]
"""

from __future__ import annotations

import os
import re
import sys
import shutil
import argparse
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

_SCRIPT_DIR = os.path.abspath(os.path.dirname(__file__))

DEFAULT_ROOT    = r"C:\Users\tanne\Downloads\SC_TEB\2025\extracted"
DEFAULT_OUT_DIR = os.path.join(_SCRIPT_DIR, "output")
DEFAULT_LOG_DIR = os.path.join(_SCRIPT_DIR, "logs")

# ---------------------------------------------------------------------------
# PDF ENGINE
# ---------------------------------------------------------------------------

try:
    import fitz as _fitz
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_DATA_FONT   = Font(name="Arial", size=10)
_ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
_OK_FILL     = PatternFill("solid", start_color="C6EFCE")
_ERR_FILL    = PatternFill("solid", start_color="FFC7CE")
_BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _hrow(ws, row, labels):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font  = _HEADER_FONT
        c.fill  = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _drow(ws, row, values, alt=False, status_col=None):
    fill = _ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = _DATA_FONT
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if status_col and col == status_col:
            c.fill = _OK_FILL if val == "OK" else _ERR_FILL
        elif fill:
            c.fill = fill


def _auto_width(ws, max_w=60):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 4, max_w)


# ---------------------------------------------------------------------------
# PDF DISCOVERY  (identical to p01 / p03)
# ---------------------------------------------------------------------------

def _find_pdfs(start_dir: str) -> list[str]:
    pdfs = []
    for root, _, files in os.walk(start_dir):
        for fn in files:
            if fn.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, fn))
    return pdfs


def _find_tab1_dir(project_root: str) -> str | None:
    tab1_candidates, tab3_candidates = [], []
    for dirpath, dirnames, _ in os.walk(project_root):
        for dname in dirnames:
            name = dname.lower()
            full = os.path.join(dirpath, dname)
            if re.search(r"\btab\s*0?1\b", name):
                tab1_candidates.append(full)
            elif name.startswith("1 ") and re.search(r"app", name):
                tab1_candidates.append(full)
            elif re.search(r"\btab\b", name) and re.search(r"\b3\b", name):
                tab3_candidates.append(full)
            elif name.startswith("tab 3") or name.startswith("tab3") or "tab 03" in name:
                tab3_candidates.append(full)
    candidates = tab1_candidates + tab3_candidates
    if not candidates:
        return None
    candidates.sort(key=lambda p: len(p.split(os.sep)))
    return candidates[0]


def _score_pdf(pdf_path: str) -> int:
    score = 0
    name = os.path.basename(pdf_path).lower().strip()
    bad = [
        "waiver", "map", "railroad", "site plan", "zoning", "utility",
        "environmental", "plans", "specifications", "appraisal",
        "market study", "syndication", "architect", "engineer",
        "certification", "site control", "checklist", "entity",
        "agreement", "opinion", "survey", "title", "easement",
        "phase i", "phase 1", "soil", "geotech",
    ]
    for w in bad:
        if w in name: score -= 12
    if "executed" in name and not re.search(r"teb.{0,20}app", name):
        score -= 25
    if "application page" in name: score -= 15
    if "application" in name: score += 12
    elif "app" in name: score += 6
    if "teb" in name: score += 10
    if "teb" in name and ("app" in name or "application" in name): score += 20
    if name in ("2025 teb application.pdf", "2026 teb application.pdf"): score += 30
    if re.search(r"teb.?app", name): score += 25
    try:
        size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        score += min(int(size_mb), 20)
        if size_mb < 1.0: score -= 10
    except Exception:
        pass
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:15]:
                txt = page.extract_text() or ""
                if re.search(r"Proforma Income Statement", txt, re.I) and \
                   re.search(r"Net Operating Income", txt, re.I):
                    score += 60
                    break
    except Exception:
        pass
    return score


def pick_best_pdf(project_path: str) -> str | None:
    tab1 = _find_tab1_dir(project_path)
    pdfs = _find_pdfs(tab1) if tab1 else []
    if not pdfs:
        pdfs = _find_pdfs(project_path)
    if not pdfs:
        return None
    scored = sorted(
        [((_score_pdf(p), os.path.getsize(p) if os.path.exists(p) else 0), p) for p in pdfs],
        reverse=True,
    )
    return scored[0][1]


# ---------------------------------------------------------------------------
# PAGE 7 DETECTION
# ---------------------------------------------------------------------------

def _extract_all_pages_text(pdf_path: str) -> list[str]:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        if any(len(t) > 200 for t in pages):
            return pages
    except Exception:
        pages = []
    if _FITZ_AVAILABLE:
        try:
            doc = _fitz.open(pdf_path)
            fitz_pages = [page.get_text("text") or "" for page in doc]
            doc.close()
            if any(len(t) > 100 for t in fitz_pages):
                return fitz_pages
        except Exception:
            pass
    return pages


def _score_page7(txt: str) -> int:
    score = 0
    if re.search(r"Proforma Income Statement",        txt, re.I): score += 40
    if re.search(r"Net Operating Income",             txt, re.I): score += 30
    if re.search(r"Effective Gross Income",           txt, re.I): score += 25
    if re.search(r"Total Administrative",             txt, re.I): score += 20
    if re.search(r"Vacancy Allowance",                txt, re.I): score += 15
    if re.search(r"Page\s*7\b",                       txt, re.I): score += 10
    return score


def _find_page7_text(pdf_path: str) -> str | None:
    pages = _extract_all_pages_text(pdf_path)
    best_score, best_text = 0, None
    for txt in pages:
        score = _score_page7(txt)
        if score > best_score:
            best_score = score
            best_text  = txt
    return best_text if best_score >= 40 else None


# ---------------------------------------------------------------------------
# FIELD HELPERS
# ---------------------------------------------------------------------------

def _dollar(label: str, txt: str) -> str:
    """
    Extract dollar amount for a labeled line item.
    Handles two-column layout (amount followed by next column label),
    end-of-line amounts, and dash (zero).
    """
    escaped = re.escape(label)
    # Mid-line: number followed by a capital letter (next column)
    m = re.search(escaped + r"\s+([\d,]+\.?\d*)\s+[A-Z]", txt, re.I | re.M)
    if m: return m.group(1).replace(",", "")
    # End of line
    m = re.search(escaped + r"\s+([\d,]+\.?\d*)\s*$", txt, re.I | re.M)
    if m: return m.group(1).replace(",", "")
    # Dash = zero (mid-line or EOL)
    m = re.search(escaped + r"\s+-\s*(?:[A-Z\n]|$)", txt, re.I | re.M)
    if m: return "0"
    return ""


def _paren(label: str, txt: str) -> str:
    """Extract parenthesized negative amount."""
    m = re.search(re.escape(label) + r".*?\(([\d,]+\.?\d*)\)", txt, re.I)
    return "-" + m.group(1).replace(",", "") if m else ""


def _pct(label: str, txt: str) -> str:
    m = re.search(re.escape(label) + r"\s*([\d.]+)%", txt, re.I)
    return m.group(1) if m else ""


def _x(pattern: str, txt: str) -> str:
    m = re.search(pattern, txt, re.I)
    return m.group(1).strip().replace(",", "") if m else ""


# ---------------------------------------------------------------------------
# PARSE PAGE 7
# ---------------------------------------------------------------------------

def parse_page7(pdf_path: str, project: str) -> dict:
    try:
        txt = _find_page7_text(pdf_path)
        if txt is None:
            return {"status": "not_found", "row": {}}
    except Exception as e:
        return {"status": f"error: {e}", "row": {}}

    row = {
        "project_folder": project,
        "pdf_path":        pdf_path,

        # Rental income
        "li_rental_income":       _dollar("From Low Income Units", txt),
        "mr_rental_income":       _dollar("From Market Rate Units", txt),
        "total_rental_income":    _dollar("Total Annual Rental Income", txt),
        "other_income":           _dollar("Other Income", txt),
        "vacancy_pct":            _pct("*Vacancy%", txt),
        "vacancy_allowance":      _paren("Vacancy Allowance =", txt),
        "egi":                    _x(r"Effective Gross Income \(EGI\)\s*=\s*([\d,]+\.?\d*)", txt),

        # Administrative expenses
        "accounting_audit":       _dollar("Accounting/Audit", txt),
        "advertising":            _dollar("Advertising", txt),
        "compliance_fees":        _dollar("Annual Compliance Fees", txt),
        "legal":                  _dollar("Legal", txt),
        "licenses_permits":       _dollar("Licenses and Permits", txt),
        "management_fees":        _dollar("Management Fees", txt),
        "management_payroll":     _dollar("Management Payroll", txt),
        "mgmt_payroll_taxes":     _dollar("Management Payroll Taxes", txt),
        "telephone":              _dollar("Telephone", txt),
        "office_supplies":        _dollar("Office Supplies", txt),
        "other_admin":            _dollar("Other Admin. Expenses (7-A)", txt),
        "total_admin":            _dollar("Total Administrative", txt),
        "admin_pct_egi":          _pct("Percent of EGI", txt),

        # Maintenance expenses
        "clubhouse_maint":        _dollar("Clubhouse Maintenance", txt),
        "decorating":             _dollar("Decorating", txt),
        "elevator":               _dollar("Elevator", txt),
        "extermination":          _dollar("Extermination", txt),
        "landscaping":            _dollar("Landscaping", txt),
        "maint_payroll":          _dollar("Maintenance Payroll", txt),
        "maint_payroll_taxes":    _dollar("Maintenance Payroll Taxes", txt),
        "parking_lot":            _dollar("Parking Lot Maintenance", txt),
        "repairs":                _dollar("Repairs", txt),
        "supplies":               _dollar("Supplies", txt),
        "pool_maint":             _dollar("Pool Maintenance", txt),
        "other_maint":            _dollar("Other Maintenance (7-A)", txt),
        "total_maint":            _dollar("Total Maintenance", txt),

        # Operating expenses
        "fuel":                   _dollar("Fuel", txt),
        "electrical":             _dollar("Electrical", txt),
        "water_sewer":            _dollar("Water and Sewer", txt),
        "natural_gas":            _dollar("Natural gas", txt),
        "trash":                  _dollar("Trash", txt),
        "security":               _dollar("Security", txt),
        "other_operating":        _dollar("Other Operating (7-A)", txt),
        "total_operating":        _dollar("Total Operating", txt),

        # Fixed expenses
        "insurance":              _dollar("Insurance", txt),
        "real_estate_taxes":      _dollar("Real Estate Taxes", txt),
        "other_taxes":            _dollar("Other Taxes (7-A)", txt),
        "total_fixed":            _dollar("Total Fixed Expenses", txt),

        # Totals & reserves
        "total_annual_expenses":  _dollar("Total Annual Expenses", txt),
        "replacement_reserves":   _dollar("Replacement Reserves", txt),
        "total_reserves":         _dollar("Total Reserves", txt),
        "noi":                    _dollar("Net Operating Income", txt),
        "other_income_pct":       _x(r"Other Income / Rental Income\s*=\s*([\d.]+)%", txt),

        "status": "OK",
    }

    return {"status": "OK", "row": row}


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

COLUMNS = [
    ("project_folder",       "Project Folder"),
    # Income
    ("li_rental_income",     "LI Rental Income"),
    ("mr_rental_income",     "MR Rental Income"),
    ("total_rental_income",  "Total Rental Income"),
    ("other_income",         "Other Income"),
    ("vacancy_pct",          "Vacancy %"),
    ("vacancy_allowance",    "Vacancy Allowance"),
    ("egi",                  "EGI"),
    # Admin
    ("accounting_audit",     "Accounting/Audit"),
    ("advertising",          "Advertising"),
    ("compliance_fees",      "Compliance Fees"),
    ("legal",                "Legal"),
    ("licenses_permits",     "Licenses & Permits"),
    ("management_fees",      "Management Fees"),
    ("management_payroll",   "Management Payroll"),
    ("mgmt_payroll_taxes",   "Mgmt Payroll Taxes"),
    ("telephone",            "Telephone"),
    ("office_supplies",      "Office Supplies"),
    ("other_admin",          "Other Admin (7-A)"),
    ("total_admin",          "Total Administrative"),
    ("admin_pct_egi",        "Admin % EGI"),
    # Maintenance
    ("clubhouse_maint",      "Clubhouse Maint."),
    ("decorating",           "Decorating"),
    ("elevator",             "Elevator"),
    ("extermination",        "Extermination"),
    ("landscaping",          "Landscaping"),
    ("maint_payroll",        "Maint. Payroll"),
    ("maint_payroll_taxes",  "Maint. Payroll Taxes"),
    ("parking_lot",          "Parking Lot Maint."),
    ("repairs",              "Repairs"),
    ("supplies",             "Supplies"),
    ("pool_maint",           "Pool Maint."),
    ("other_maint",          "Other Maint. (7-A)"),
    ("total_maint",          "Total Maintenance"),
    # Operating
    ("fuel",                 "Fuel"),
    ("electrical",           "Electrical"),
    ("water_sewer",          "Water & Sewer"),
    ("natural_gas",          "Natural Gas"),
    ("trash",                "Trash"),
    ("security",             "Security"),
    ("other_operating",      "Other Operating (7-A)"),
    ("total_operating",      "Total Operating"),
    # Fixed
    ("insurance",            "Insurance"),
    ("real_estate_taxes",    "Real Estate Taxes"),
    ("other_taxes",          "Other Taxes (7-A)"),
    ("total_fixed",          "Total Fixed Expenses"),
    # Summary
    ("total_annual_expenses","Total Annual Expenses"),
    ("replacement_reserves", "Replacement Reserves"),
    ("total_reserves",       "Total Reserves"),
    ("noi",                  "NOI"),
    ("other_income_pct",     "Other Income / Rental Income %"),
    ("status",               "Status"),
    ("pdf_path",             "PDF Path"),
]


def build_excel(rows: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma"
    ws.freeze_panes = "A2"

    keys       = [k for k, _ in COLUMNS]
    labels     = [l for _, l in COLUMNS]
    status_col = keys.index("status") + 1

    _hrow(ws, 1, labels)
    for i, row in enumerate(rows):
        _drow(ws, i + 2, [row.get(k, "") for k in keys],
              alt=(i % 2 == 1), status_col=status_col)

    _auto_width(ws)
    wb.save(output_path)
    print("Excel written to: " + output_path)


# ---------------------------------------------------------------------------
# CLI + MAIN
# ---------------------------------------------------------------------------

def _parse_args():
    p = argparse.ArgumentParser(description="Extract Page 7 proforma from SC TEB PDFs.")
    p.add_argument("--root",       default=DEFAULT_ROOT,    help="Root folder of project subfolders.")
    p.add_argument("--output-dir", default=DEFAULT_OUT_DIR, help="Output directory.")
    p.add_argument("--log-dir",    default=DEFAULT_LOG_DIR, help="Log directory.")
    return p.parse_args()


def main():
    args = _parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    os.makedirs(args.log_dir,    exist_ok=True)

    run_ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_xlsx = os.path.join(args.output_dir, f"SC_TEB_P07_Proforma_{run_ts}.xlsx")
    latest_xlsx = os.path.join(args.output_dir, "SC_TEB_P07_Proforma_latest.xlsx")
    log_path    = os.path.join(args.log_dir,    f"p07_run_{run_ts}.log")

    log_file        = open(log_path, "w", encoding="utf-8")
    original_stdout = sys.stdout

    class Tee:
        def __init__(self, *s): self.streams = s
        def write(self, d):
            for s in self.streams: s.write(d); s.flush()
        def flush(self):
            for s in self.streams: s.flush()

    sys.stdout = Tee(original_stdout, log_file)

    try:
        print("Root:       " + args.root)
        print("Output dir: " + args.output_dir)
        print()

        if not os.path.isdir(args.root):
            raise FileNotFoundError('Root folder not found: "' + args.root + '"')

        rows = []

        for project in sorted(os.listdir(args.root)):
            project_path = os.path.join(args.root, project)
            if not os.path.isdir(project_path):
                continue

            pdf_path = pick_best_pdf(project_path)
            if not pdf_path:
                print("[SKIP] " + project + " – no PDF found")
                rows.append({"project_folder": project, "status": "No PDF found"})
                continue

            result = parse_page7(pdf_path, project)

            if result["status"] == "OK":
                row = result["row"]
                rows.append(row)
                print("[OK]   " + project +
                      " – NOI=" + row.get("noi", "") +
                      " EGI=" + row.get("egi", "") +
                      " vacancy=" + row.get("vacancy_pct", "") + "%")
            else:
                rows.append({"project_folder": project, "pdf_path": pdf_path,
                             "status": result["status"]})
                print("[MISS] " + project + " – " + result["status"])

        build_excel(rows, output_xlsx)
        shutil.copy2(output_xlsx, latest_xlsx)

        print("\nProjects processed: " + str(len(rows)))
        print("Excel:  " + output_xlsx)
        print("Latest: " + latest_xlsx)
        print("Log:    " + log_path)

    finally:
        sys.stdout = original_stdout
        log_file.close()


if __name__ == "__main__":
    main()
