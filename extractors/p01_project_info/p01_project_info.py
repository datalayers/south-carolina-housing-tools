"""
p01_project_info.py — SC TEB Application Page 1 Extractor
==========================================================
Extracts project information from Page 1 of each SC TEB application PDF.

Output: single Excel workbook with one sheet — Project_Info (one row per project)

Usage:
    python p01_project_info.py [--root ROOT] [--output-dir DIR] [--log-dir DIR]
"""

from __future__ import annotations

import os
import re
import sys
import argparse
import shutil
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

_SCRIPT_DIR = os.path.abspath(os.path.dirname(__file__))

DEFAULT_ROOT    = r"C:\Users\tanne\Downloads\SC_TEB\2025\extracted"
DEFAULT_OUT_DIR = os.path.join(_SCRIPT_DIR, "output")
DEFAULT_LOG_DIR = os.path.join(_SCRIPT_DIR, "logs")

# ---------------------------------------------------------------------------
# PDF ENGINE
# ---------------------------------------------------------------------------

try:
    import fitz as _fitz
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_DATA_FONT   = Font(name="Arial", size=10)
_ALT_FILL    = PatternFill("solid", start_color="DCE6F1")
_OK_FILL     = PatternFill("solid", start_color="C6EFCE")
_ERR_FILL    = PatternFill("solid", start_color="FFC7CE")
_BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _hrow(ws, row, labels):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font  = _HEADER_FONT
        c.fill  = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _drow(ws, row, values, alt=False, status_col=None):
    fill = _ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = _DATA_FONT
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if status_col and col == status_col:
            c.fill = _OK_FILL if val == "OK" else _ERR_FILL
        elif fill:
            c.fill = fill


def _auto_width(ws, max_w=60):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 4, max_w)


# ---------------------------------------------------------------------------
# PDF DISCOVERY  (identical logic to p03_site_info)
# ---------------------------------------------------------------------------

def _find_pdfs(start_dir: str) -> list[str]:
    pdfs = []
    for root, _, files in os.walk(start_dir):
        for fn in files:
            if fn.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, fn))
    return pdfs


def _find_tab1_dir(project_root: str) -> str | None:
    tab1_candidates = []
    tab3_candidates = []
    for dirpath, dirnames, _ in os.walk(project_root):
        for dname in dirnames:
            name = dname.lower()
            full = os.path.join(dirpath, dname)
            if re.search(r"\btab\s*0?1\b", name):
                tab1_candidates.append(full)
            elif name.startswith("1 ") and re.search(r"app", name):
                tab1_candidates.append(full)
            elif re.search(r"\btab\b", name) and re.search(r"\b3\b", name):
                tab3_candidates.append(full)
            elif name.startswith("tab 3") or name.startswith("tab3") or "tab 03" in name:
                tab3_candidates.append(full)
    candidates = tab1_candidates + tab3_candidates
    if not candidates:
        return None
    candidates.sort(key=lambda p: len(p.split(os.sep)))
    return candidates[0]


def _score_pdf(pdf_path: str) -> int:
    score = 0
    name = os.path.basename(pdf_path).lower().strip()
    bad = [
        "waiver", "map", "railroad", "site plan", "zoning", "utility",
        "environmental", "plans", "specifications", "appraisal",
        "market study", "syndication", "architect", "engineer",
        "certification", "site control", "checklist", "entity",
        "agreement", "opinion", "survey", "title", "easement",
        "phase i", "phase 1", "soil", "geotech",
    ]
    for w in bad:
        if w in name: score -= 12
    if "executed" in name and not re.search(r"teb.{0,20}app", name):
        score -= 25
    if "application page" in name:
        score -= 15
    if "application" in name:
        score += 12
    elif "app" in name:
        score += 6
    if "teb" in name:
        score += 10
    if "teb" in name and ("app" in name or "application" in name):
        score += 20
    if name in ("2025 teb application.pdf", "2026 teb application.pdf"):
        score += 30
    if re.search(r"teb.?app", name):
        score += 25
    try:
        size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        score += min(int(size_mb), 20)
        if size_mb < 1.0:
            score -= 10
    except Exception:
        pass
    # Content check: prefer PDFs containing Page 1 anchor text
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:15]:
                txt = page.extract_text() or ""
                if re.search(r"Applicant Information", txt, re.I) and \
                   re.search(r"Contact Person", txt, re.I):
                    score += 60
                    break
    except Exception:
        pass
    return score


def pick_best_pdf(project_path: str) -> str | None:
    tab1 = _find_tab1_dir(project_path)
    pdfs = _find_pdfs(tab1) if tab1 else []
    if not pdfs:
        pdfs = _find_pdfs(project_path)
    if not pdfs:
        return None
    scored = sorted(
        [((_score_pdf(p), os.path.getsize(p) if os.path.exists(p) else 0), p) for p in pdfs],
        reverse=True,
    )
    return scored[0][1]


# ---------------------------------------------------------------------------
# PAGE 1 DETECTION
# ---------------------------------------------------------------------------

def _extract_all_pages_text(pdf_path: str) -> list[str]:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        if any(len(t) > 200 for t in pages):
            return pages
    except Exception:
        pages = []
    if _FITZ_AVAILABLE:
        try:
            doc = _fitz.open(pdf_path)
            fitz_pages = [page.get_text("text") or "" for page in doc]
            doc.close()
            if any(len(t) > 100 for t in fitz_pages):
                return fitz_pages
        except Exception:
            pass
    return pages


def _score_page1(txt: str) -> int:
    score = 0
    if re.search(r"Applicant Information", txt, re.I):       score += 30
    if re.search(r"Contact Person", txt, re.I):              score += 25
    if re.search(r"Total # of Low-Income Units", txt, re.I): score += 25
    if re.search(r"Application Type", txt, re.I):            score += 20
    if re.search(r"Entity Name", txt, re.I):                 score += 15
    if re.search(r"Page\s*1\b", txt, re.I):                  score += 10
    if re.search(r"Fed ID #", txt, re.I):                    score += 15
    return score


def _find_page1_text(pdf_path: str) -> str | None:
    pages = _extract_all_pages_text(pdf_path)
    best_score, best_text = 0, None
    for txt in pages:
        score = _score_page1(txt)
        if score > best_score:
            best_score = score
            best_text  = txt
    return best_text if best_score >= 30 else None


# ---------------------------------------------------------------------------
# FIELD HELPERS
# ---------------------------------------------------------------------------

def _x(pattern: str, txt: str, group: int = 1) -> str:
    m = re.search(pattern, txt, re.I | re.DOTALL)
    return re.sub(r"\s+", " ", m.group(group)).strip() if m else ""


def _checked(label: str, txt: str) -> str:
    """Return Y if label has X/x before or after it."""
    m = re.search(r"(X|x)\s*" + re.escape(label) + r"|" + re.escape(label) + r"\s*(X|x)", txt, re.I)
    return "Y" if m else "N"


def _entity_type(txt: str) -> str:
    for etype in ["Limited Partnership", "Limited Liability Company", "Non-Profit"]:
        if re.search(r"(X|x)\s*" + re.escape(etype), txt, re.I):
            return etype
    return ""


# ---------------------------------------------------------------------------
# PARSE PAGE 1
# ---------------------------------------------------------------------------

def parse_page1(pdf_path: str, project: str) -> dict:
    try:
        txt = _find_page1_text(pdf_path)
        if txt is None:
            return {"status": "not_found", "row": {}}
    except Exception as e:
        return {"status": f"error: {e}", "row": {}}

    row = {
        "project_folder":       project,
        "pdf_path":             pdf_path,

        # Identity
        "development_name":     _x(r"Development Name:\s*(.+?)\s+(?:County:|Date:)", txt),
        "application_date":     _x(r"Date:\s*(\d{1,2}/\d{1,2}/\d{4})", txt),
        "application_type":     _x(r"Application Type:\s*(.+?)(?:\n|$)", txt),

        # Credit types
        "credit_9pct":          _checked("9% Tax Credit", txt),
        "credit_4pct":          _checked("4% Tax Credit", txt),
        "state_tax_credits":    _checked("State Tax Credits", txt),

        # Application types
        "new_construction":     _checked("New Construction", txt),
        "rehabilitation":       _checked("Rehabilitation", txt),
        "acq_rehabilitation":   _checked("Acq/Rehabilitation", txt),
        "public_housing_auth":  _checked("Public Housing Authority", txt),
        "adaptive_reuse":       _checked("Adaptive Reuse", txt),

        # Unit counts
        "li_units":             _x(r"Total # of Low-Income Units:\s*(\d+)", txt),
        "mr_units":             _x(r"Total # Market Rate Units:\s*(\d+)", txt),
        "total_units":          _x(r"Total # of Units:\s*(\d+)", txt),
        "employee_units":       _x(r"Employee Units:\s*(\d+)", txt),
        "families_units":       _x(r"Designed for Families Units:\s*(\d+)", txt),
        "older_55_units":       _x(r"Older Persons \(55\+\) Units:\s*(\d+)", txt),
        "elderly_62_units":     _x(r"Elderly Persons \(62\+\) Units:\s*(\d+)", txt),
        "transitional_units":   _x(r"Transitional Units\s*(\d+)", txt),
        "homeless_units":       _x(r"Homeless Units\s*(\d+)", txt),
        "sro_units":            _x(r"Single Room Occupancy\s*(\d+)", txt),
        "supportive_units":     _x(r"Supportive Housing Units:\s*(\d+)", txt),
        "three_br_plus_units":  _x(r"3\+ Bedroom Units:\s*(\d+)", txt),

        # Site / location
        "county":               _x(r"County:\s*([A-Za-z\s]+?)\s+(?:Group:|County Code:)", txt),
        "group":                _x(r"Group:\s*([A-Z]?)\s", txt),
        "county_code":          _x(r"County Code:\s*(\d+)", txt),
        "street_address":       _x(r"Street Address:\s*(.+?)\s+County Code:", txt),
        "city":                 _x(r"City:\s*([A-Za-z\s]+?)\s+Congressional", txt),
        "state":                "SC",
        "zip":                  _x(r"Zip:\s*(\d{5})\s+Est\.", txt),
        "est_start_date":       _x(r"Est\. Start Date:\s*(\d{1,2}/\d{1,2}/\d{4})", txt),
        "congressional_district": _x(r"Congressional District #\s*:\s*(\d+)", txt),

        # Entity
        "entity_type":          _entity_type(txt),
        "entity_name":          _x(r"Entity Name:\s*(.+?)(?:\n|Street Address:)", txt),
        "entity_street":        _x(r"Limited Liability Company Street Address:\s*(.+?)(?:\n|City:)", txt),
        "entity_city":          _x(r"Non-Profit City:\s*(.+?)\s+State:", txt),
        "entity_state":         _x(r"Non-Profit City:.*?State:\s*([A-Za-z\s]+?)\s+Zip:", txt),
        "entity_zip":           _x(r"Non-Profit City:.*?Zip:\s*(\d{5})", txt),
        "fed_id":               _x(r"Fed ID #\s*:\s*([\d-]+)", txt),

        # Contact
        "contact_person":       _x(r"Contact Person:\s*(.+?)\s+Telephone:", txt),
        "telephone":            _x(r"Telephone:\s*([\d\s\-().]+?)(?:\n|Email:)", txt),
        "email":                _x(r"Email:\s*(\S+@\S+)", txt),

        # Principals
        "num_applications":     _x(r"associated with\?\s*(\d+)", txt),

        "status":               "OK",
    }

    return {"status": "OK", "row": row}


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

COLUMNS = [
    ("project_folder",          "Project Folder"),
    ("development_name",        "Development Name"),
    ("application_date",        "Application Date"),
    ("application_type",        "Application Type"),
    ("credit_9pct",             "9% Tax Credit"),
    ("credit_4pct",             "4% Tax Credit"),
    ("state_tax_credits",       "State Tax Credits"),
    ("new_construction",        "New Construction"),
    ("rehabilitation",          "Rehabilitation"),
    ("acq_rehabilitation",      "Acq/Rehabilitation"),
    ("public_housing_auth",     "Public Housing Auth"),
    ("adaptive_reuse",          "Adaptive Reuse"),
    ("li_units",                "LI Units"),
    ("mr_units",                "MR Units"),
    ("total_units",             "Total Units"),
    ("employee_units",          "Employee Units"),
    ("families_units",          "Families Units"),
    ("older_55_units",          "Older 55+ Units"),
    ("elderly_62_units",        "Elderly 62+ Units"),
    ("transitional_units",      "Transitional Units"),
    ("homeless_units",          "Homeless Units"),
    ("sro_units",               "SRO Units"),
    ("supportive_units",        "Supportive Units"),
    ("three_br_plus_units",     "3+ BR Units"),
    ("county",                  "County"),
    ("group",                   "Group"),
    ("county_code",             "County Code"),
    ("street_address",          "Street Address"),
    ("city",                    "City"),
    ("state",                   "State"),
    ("zip",                     "Zip"),
    ("est_start_date",          "Est. Start Date"),
    ("congressional_district",  "Congressional District"),
    ("entity_type",             "Entity Type"),
    ("entity_name",             "Entity Name"),
    ("entity_street",           "Entity Street"),
    ("entity_city",             "Entity City"),
    ("entity_state",            "Entity State"),
    ("entity_zip",              "Entity Zip"),
    ("fed_id",                  "Fed ID #"),
    ("contact_person",          "Contact Person"),
    ("telephone",               "Telephone"),
    ("email",                   "Email"),
    ("num_applications",        "# Applications by Principals"),
    ("status",                  "Status"),
    ("pdf_path",                "PDF Path"),
]


def build_excel(rows: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Project_Info"
    ws.freeze_panes = "A2"

    keys   = [k for k, _ in COLUMNS]
    labels = [l for _, l in COLUMNS]
    status_col = keys.index("status") + 1

    _hrow(ws, 1, labels)
    for i, row in enumerate(rows):
        _drow(ws, i + 2, [row.get(k, "") for k in keys],
              alt=(i % 2 == 1), status_col=status_col)

    _auto_width(ws)
    wb.save(output_path)
    print("Excel written to: " + output_path)


# ---------------------------------------------------------------------------
# CLI + MAIN
# ---------------------------------------------------------------------------

def _parse_args():
    p = argparse.ArgumentParser(description="Extract Page 1 project info from SC TEB PDFs.")
    p.add_argument("--root",       default=DEFAULT_ROOT,    help="Root folder of project subfolders.")
    p.add_argument("--output-dir", default=DEFAULT_OUT_DIR, help="Output directory.")
    p.add_argument("--log-dir",    default=DEFAULT_LOG_DIR, help="Log directory.")
    return p.parse_args()


def main():
    args = _parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    os.makedirs(args.log_dir,    exist_ok=True)

    run_ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_xlsx = os.path.join(args.output_dir, f"SC_TEB_P01_ProjectInfo_{run_ts}.xlsx")
    latest_xlsx = os.path.join(args.output_dir, "SC_TEB_P01_ProjectInfo_latest.xlsx")
    log_path    = os.path.join(args.log_dir,    f"p01_run_{run_ts}.log")

    log_file        = open(log_path, "w", encoding="utf-8")
    original_stdout = sys.stdout

    class Tee:
        def __init__(self, *s): self.streams = s
        def write(self, d):
            for s in self.streams: s.write(d); s.flush()
        def flush(self):
            for s in self.streams: s.flush()

    sys.stdout = Tee(original_stdout, log_file)

    try:
        print("Root:       " + args.root)
        print("Output dir: " + args.output_dir)
        print()

        if not os.path.isdir(args.root):
            raise FileNotFoundError('Root folder not found: "' + args.root + '"')

        rows = []

        for project in sorted(os.listdir(args.root)):
            project_path = os.path.join(args.root, project)
            if not os.path.isdir(project_path):
                continue

            pdf_path = pick_best_pdf(project_path)
            if not pdf_path:
                print("[SKIP] " + project + " – no PDF found")
                rows.append({"project_folder": project, "status": "No PDF found"})
                continue

            result = parse_page1(pdf_path, project)

            if result["status"] == "OK":
                row = result["row"]
                rows.append(row)
                print("[OK]   " + project + " – " + row.get("development_name", "") +
                      " | " + row.get("county", "") + " | units=" + row.get("total_units", ""))
            else:
                rows.append({"project_folder": project, "pdf_path": pdf_path,
                             "status": result["status"]})
                print("[MISS] " + project + " – " + result["status"])

        build_excel(rows, output_xlsx)
        shutil.copy2(output_xlsx, latest_xlsx)

        print("\nProjects processed: " + str(len(rows)))
        print("Excel:  " + output_xlsx)
        print("Latest: " + latest_xlsx)
        print("Log:    " + log_path)

    finally:
        sys.stdout = original_stdout
        log_file.close()


if __name__ == "__main__":
    main()
