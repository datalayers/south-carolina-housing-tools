"""
p03_site_info.py — SC TEB Application Page 3 Extractor
=======================================================
Extracts site information from Page 3 of each SC TEB application PDF:

  Sheet 1 — Site_Summary   : one row per project (flags, districts, coordinates)
  Sheet 2 — Site_Parcels   : one row per parcel (up to 3 per project)

Usage:
    python p03_site_info.py [--root ROOT] [--output-dir DIR] [--log-dir DIR]
"""

from __future__ import annotations

import os
import re
import sys
import argparse
from datetime import datetime

import pdfplumber
try:
    import fitz as _fitz  # PyMuPDF — primary text engine
    _FITZ_AVAILABLE = True
except ImportError:
    _FITZ_AVAILABLE = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

_SCRIPT_DIR = os.path.abspath(os.path.dirname(__file__))

DEFAULT_ROOT    = r"C:\Users\tanne\Downloads\SC_TEB\2025\extracted"
DEFAULT_OUT_DIR = os.path.join(_SCRIPT_DIR, "output")
DEFAULT_LOG_DIR = os.path.join(_SCRIPT_DIR, "logs")

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------

_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
_DATA_FONT    = Font(name="Arial", size=10)
_ALT_FILL     = PatternFill("solid", start_color="DCE6F1")
_SECTION_FONT = Font(name="Arial", bold=True, size=10, color="1F4E79")
_SECTION_FILL = PatternFill("solid", start_color="D9E1F2")
_OK_FILL      = PatternFill("solid", start_color="C6EFCE")
_ERR_FILL     = PatternFill("solid", start_color="FFC7CE")
_BORDER       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _hrow(ws, row, labels):
    for col, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _drow(ws, row, values, alt=False, status_col=None):
    fill = _ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _DATA_FONT
        c.border = _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if status_col and col == status_col:
            c.fill = _OK_FILL if val == "OK" else _ERR_FILL
        elif fill:
            c.fill = fill


def _auto_width(ws, max_w=60):
    for col in ws.columns:
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 4, max_w)


# ---------------------------------------------------------------------------
# PDF DISCOVERY  (same pattern as other extractors)
# ---------------------------------------------------------------------------

def _find_pdfs(start_dir: str) -> list[str]:
    pdfs = []
    for root, _, files in os.walk(start_dir):
        for fn in files:
            if fn.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, fn))
    return pdfs


def _find_tab3_dir(project_root: str) -> str | None:
    """
    Page 3 lives inside the main TEB application PDF, which is in TAB 1.
    Search for TAB 1 - Application first, then TAB 3 as a fallback.
    """
    tab1_candidates = []
    tab3_candidates = []
    for dirpath, dirnames, _ in os.walk(project_root):
        for dname in dirnames:
            name = dname.lower()
            full = os.path.join(dirpath, dname)
            # Match "TAB 1", "TAB 01", "Tab 1", "TAB1", "TAB01" with or without "app"
            if re.search(r"\btab\s*0?1\b", name):
                tab1_candidates.append(full)
            elif name.startswith("1 ") and re.search(r"app", name):
                tab1_candidates.append(full)
            elif re.search(r"\btab\b", name) and re.search(r"\b3\b", name):
                tab3_candidates.append(full)
            elif name.startswith("tab 3") or name.startswith("tab3") or "tab 03" in name:
                tab3_candidates.append(full)
    # TAB 1 takes priority since page 3 is in the main application PDF
    candidates = tab1_candidates + tab3_candidates
    if not candidates:
        return None
    candidates.sort(key=lambda p: len(p.split(os.sep)))
    return candidates[0]


def _score_pdf(pdf_path: str) -> int:
    score = 0
    name = os.path.basename(pdf_path).lower().strip()
    bad = [
        "waiver", "map", "railroad", "site plan", "zoning", "utility",
        "environmental", "plans", "specifications", "appraisal",
        "market study", "syndication", "architect", "engineer",
        "certification", "site control", "checklist", "entity",
        "agreement", "opinion", "survey", "title", "easement",
        "phase i", "phase 1", "soil", "geotech",
    ]
    for w in bad:
        if w in name:
            score -= 12
    # "executed" penalizes attachments but not the TEB application itself
    if "executed" in name and not re.search(r"teb.{0,20}app", name):
        score -= 25
    if "application page" in name:
        score -= 15
    if "application" in name:
        score += 12
    elif "app" in name:
        score += 6
    if "teb" in name:
        score += 10
    if "teb" in name and ("app" in name or "application" in name):
        score += 20
    if name in ("2025 teb application.pdf", "2026 teb application.pdf"):
        score += 30
    if re.search(r"teb.?app", name):
        score += 25  # e.g. "2025-4-TEB-App", "TEB Application - Signed"
    try:
        size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
        score += min(int(size_mb), 20)
        if size_mb < 1.0:
            score -= 10
    except Exception:
        pass
    # Content check: strongly prefer PDFs that contain the Page 3 anchor
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:15]:
                txt = page.extract_text() or ""
                if re.search(r"Coordinates for development centroid", txt, re.I):
                    score += 60
                    break
                if re.search(r"Site Control \(Parcel 1\)", txt, re.I):
                    score += 30
                    break
    except Exception:
        pass
    return score


def pick_best_pdf(project_path: str) -> str | None:
    tab3 = _find_tab3_dir(project_path)
    pdfs = _find_pdfs(tab3) if tab3 else []
    if not pdfs:
        pdfs = _find_pdfs(project_path)
    if not pdfs:
        return None
    scored = sorted(
        [((_score_pdf(p), os.path.getsize(p) if os.path.exists(p) else 0), p) for p in pdfs],
        reverse=True,
    )
    return scored[0][1]


# ---------------------------------------------------------------------------
# PAGE 3 DETECTION
# ---------------------------------------------------------------------------

def _score_page_text(txt: str) -> int:
    """Score a page's text for likelihood of being Page 3."""
    score = 0
    if re.search(r"Coordinates for development centroid", txt, re.I): score += 40
    if re.search(r"Site Control \(Parcel 1\)", txt, re.I):           score += 30
    if re.search(r"USDA Eligible Area", txt, re.I):                    score += 20
    if re.search(r"Located in a Flood Plain", txt, re.I):              score += 10
    if re.search(r"Page\s*3\b", txt, re.I):                          score += 15
    if re.search(r"Latitude:", txt, re.I):                             score += 20
    return score


def _extract_all_pages_text(pdf_path: str) -> list[str]:
    """
    Extract text from all pages.
    Uses pdfplumber as primary (consistent label-based text for our regexes).
    Falls back to fitz only when pdfplumber returns no meaningful text —
    this handles scanned/image PDFs that pdfplumber cannot read.
    """
    # pdfplumber primary — our regexes are tuned to its output format
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages = [page.extract_text() or "" for page in pdf.pages]
        # If pdfplumber got meaningful content on any page, use it
        if any(len(t) > 200 for t in pages):
            return pages
    except Exception:
        pages = []

    # fitz fallback — handles PDFs pdfplumber cannot read
    if _FITZ_AVAILABLE:
        try:
            doc = _fitz.open(pdf_path)
            fitz_pages = [page.get_text("text") or "" for page in doc]
            doc.close()
            if any(len(t) > 100 for t in fitz_pages):
                return fitz_pages
        except Exception:
            pass

    return pages


def _find_page3_text(pdf_path: str) -> str | None:
    """Return the text of the Page 3 content page, using best available engine."""
    pages = _extract_all_pages_text(pdf_path)
    best_score = 0
    best_text  = None
    for txt in pages:
        score = _score_page_text(txt)
        if score > best_score:
            best_score = score
            best_text  = txt
    return best_text if best_score >= 30 else None


# ---------------------------------------------------------------------------
# REGEX PATTERNS
# ---------------------------------------------------------------------------

# Y/N flags
def _yn(label: str, txt: str) -> str:
    """Extract Y or N after a labeled Y/N field. Label may contain regex patterns."""
    pattern = re.compile(label + r".*?Y/N\s+([YN])", re.I | re.DOTALL)
    m = pattern.search(txt)
    return m.group(1).upper() if m else ""


# District / tract numbers
_CONGRESSIONAL_RE  = re.compile(r"Congressional District\s*#\s*:\s*(\d+)", re.I)
_SENATE_RE         = re.compile(r"State Senate District\s*#\s*:\s*(\d+)", re.I)
_HOUSE_RE          = re.compile(r"State House District\s*#\s*:\s*(\d+)", re.I)
_CENSUS_RE         = re.compile(r"Census Tract\s*#\s*:\s*([\d.]+)", re.I)

# Coordinates — labeled style (pdfplumber) and bare style (fitz)
_LAT_RE            = re.compile(r"Latitude:\s*([-+]?\d{1,3}\.\d{4,10})", re.I)
_LON_RE            = re.compile(r"Longitude:\s*([-+]?\d{1,3}\.\d{4,10})", re.I)
_COORD_ANCHOR_RE   = re.compile(r"Coordinates\s+for\s+development\s+centroid", re.I)
_DECIMAL_RE        = re.compile(r"[-+]?\d{1,3}\.\d{4,10}")

def _sc_bounds(lat: float, lon: float) -> bool:
    return (32.0 <= lat <= 35.6) and (-84.7 <= lon <= -78.0)

def _normalize_sc(a: float, b: float):
    aa, bb = abs(a), abs(b)
    if _sc_bounds(aa, -bb): return aa, -bb
    if _sc_bounds(bb, -aa): return bb, -aa
    return None

def _extract_coords(txt: str) -> tuple:
    """Try labeled Latitude:/Longitude: first, then bare numbers near anchor.
    Always normalizes through _normalize_sc to enforce correct SC sign on longitude."""
    lat = (m := _LAT_RE.search(txt)) and m.group(1) or ""
    lon = (m := _LON_RE.search(txt)) and m.group(1) or ""
    if lat and lon:
        try:
            norm = _normalize_sc(float(lat), float(lon))
            if norm:
                return str(norm[0]), str(norm[1])
        except ValueError:
            pass
        return lat, lon  # return as-is if normalization fails
    # Fitz-style: bare numbers near the coords anchor
    m = _COORD_ANCHOR_RE.search(txt)
    if m:
        window = txt[m.start():m.end() + 300]
        nums = _DECIMAL_RE.findall(window)
        floats = []
        for n in nums:
            try: floats.append(float(n))
            except ValueError: pass
        for i in range(len(floats) - 1):
            norm = _normalize_sc(floats[i], floats[i+1])
            if norm:
                return str(norm[0]), str(norm[1])
    return "", ""

# Wetlands pct
_WETLANDS_PCT_RE   = re.compile(r"If yes, what %\?\s*([\d.]+%)", re.I)

# Parcel sections
_PARCEL_SPLIT_RE   = re.compile(
    r"Site Control \(Parcel (\d+).*?\):(.*?)(?=Site Control \(Parcel|\Z)",
    re.I | re.DOTALL
)

# Within-parcel fields
_CONTROL_RE        = re.compile(r"Control:\s*([A-Za-z /]+?)(?:\s+Expiration|\s+If Land)", re.I)
_EXPIRATION_RE     = re.compile(r"Expiration Date:\s*([\w/,.-]+?)(?:\s+If Land|\s*$)", re.I | re.M)
_LEASE_DEBT_RE     = re.compile(r"If Land Lease, how much annual debt\?\s*([\d,]+)", re.I)
_ACRES_RE          = re.compile(r"Acres:\s*([\d.,]+)", re.I)
_LAND_COST_RE      = re.compile(r"Total Cost of Land:\s*([\d,]+)", re.I)
_SELLER_RE         = re.compile(r"Seller\(s\).*?deed:\s*(.+?)(?:\n|Address:)", re.I | re.DOTALL)
_ADDRESS_RE        = re.compile(r"Address:\s*(.+?)(?:\s+City:)", re.I)
_CITY_RE           = re.compile(r"City:\s*([A-Za-z\s]+?)(?:\s+State:)", re.I)
_STATE_RE          = re.compile(r"State:\s*([A-Z]{2})", re.I)
_ZIP_RE            = re.compile(r"Zip:\s*(\d{5}(?:-\d{4})?)", re.I)
_COMMON_OWN_RE     = re.compile(r"common ownership interest.*?Y/N\s+([YN])", re.I)


def _clean(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


# ---------------------------------------------------------------------------
# PARSE PAGE 3
# ---------------------------------------------------------------------------

def parse_page3(pdf_path: str, project: str) -> dict:
    try:
        txt = _find_page3_text(pdf_path)
        if txt is None:
            return {"status": "not_found", "summary": {}, "parcels": []}
    except Exception as e:
        return {"status": f"error: {e}", "summary": {}, "parcels": []}

    # ---- Site flags ----
    summary = {
        "project_folder":              project,
        "pdf_path":                    pdf_path,
        "within_city_limits":          _yn("Development located within city limits", txt),
        "usda_eligible":               _yn("USDA Eligible Area", txt),
        "flood_plain":                 _yn("Located in a Flood Plain", txt),
        "national_register":           _yn("Listed on National Register of Historic Places", txt),
        "opportunity_zone":            _yn("Located in an Opportunity Zone", txt),
        "qualified_census_tract":      _yn("Located in a Qualified Census Tract", txt),
        "difficult_development_area":  _yn("Located in a Difficult Development Area", txt),
        "land_donated":                _yn("Was the land donated", txt),
        "site_zoned":                  _yn("Is the site zoned for your development", txt),
        "detrimental_site_characteristics": _yn("Do any detrimental site characteristics exist", txt),
        "wetlands":                    _yn("Do any wetlands.*?exist on the site", txt),
        "buildable_80pct":             _yn("Overall, is at least 80% of the site buildable", txt),
    }

    # Wetlands pct
    m = _WETLANDS_PCT_RE.search(txt)
    summary["wetlands_pct"] = m.group(1) if m else ""

    # Districts
    summary["congressional_district"] = (m := _CONGRESSIONAL_RE.search(txt)) and m.group(1) or ""
    summary["state_senate_district"]  = (m := _SENATE_RE.search(txt))         and m.group(1) or ""
    summary["state_house_district"]   = (m := _HOUSE_RE.search(txt))           and m.group(1) or ""
    summary["census_tract"]           = (m := _CENSUS_RE.search(txt))          and m.group(1) or ""

    # Coordinates (handles both pdfplumber labeled and fitz bare-number styles)
    summary["latitude"], summary["longitude"] = _extract_coords(txt)

    summary["status"] = "OK"

    # ---- Parcels ----
    parcels = []
    for pm in _PARCEL_SPLIT_RE.finditer(txt):
        parcel_num = pm.group(1)
        block      = pm.group(2)

        # Skip empty parcels (no acres, no seller, no control value)
        acres_m = _ACRES_RE.search(block)
        acres   = _clean(acres_m.group(1)) if acres_m else ""
        if not acres:
            continue

        control_m  = _CONTROL_RE.search(block)
        exp_m      = _EXPIRATION_RE.search(block)
        cost_m     = _LAND_COST_RE.search(block)
        seller_m   = _SELLER_RE.search(block)
        addr_m     = _ADDRESS_RE.search(block)
        city_m     = _CITY_RE.search(block)
        state_m    = _STATE_RE.search(block)
        zip_m      = _ZIP_RE.search(block)
        own_m      = _COMMON_OWN_RE.search(block)
        lease_m    = _LEASE_DEBT_RE.search(block)

        parcels.append({
            "project_folder":   project,
            "parcel_num":       parcel_num,
            "control_type":     _clean(control_m.group(1)) if control_m else "",
            "expiration_date":  _clean(exp_m.group(1))     if exp_m    else "",
            "acres":            acres,
            "total_cost_of_land": _clean(cost_m.group(1)) if cost_m   else "",
            "seller_name":      _clean(seller_m.group(1)) if seller_m else "",
            "address":          _clean(addr_m.group(1))   if addr_m   else "",
            "city":             _clean(city_m.group(1))   if city_m   else "",
            "state":            _clean(state_m.group(1))  if state_m  else "",
            "zip":              _clean(zip_m.group(1))     if zip_m    else "",
            "common_ownership": _clean(own_m.group(1))    if own_m    else "",
            "annual_lease_debt": _clean(lease_m.group(1)) if lease_m  else "",
        })

    return {"status": "OK", "summary": summary, "parcels": parcels}


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

SUMMARY_COLS = [
    ("project_folder",              "Project Folder"),
    ("within_city_limits",          "Within City Limits"),
    ("usda_eligible",               "USDA Eligible"),
    ("flood_plain",                 "Flood Plain"),
    ("national_register",           "National Register"),
    ("opportunity_zone",            "Opportunity Zone"),
    ("qualified_census_tract",      "Qualified Census Tract"),
    ("difficult_development_area",  "Difficult Dev Area"),
    ("land_donated",                "Land Donated"),
    ("site_zoned",                  "Site Zoned"),
    ("detrimental_site_characteristics", "Detrimental Site Char."),
    ("wetlands",                    "Wetlands"),
    ("wetlands_pct",                "Wetlands %"),
    ("buildable_80pct",             "≥80% Buildable"),
    ("congressional_district",      "Congressional District"),
    ("state_senate_district",       "State Senate District"),
    ("state_house_district",        "State House District"),
    ("census_tract",                "Census Tract"),
    ("latitude",                    "Latitude"),
    ("longitude",                   "Longitude"),
    ("status",                      "Status"),
    ("pdf_path",                    "PDF Path"),
]

PARCEL_COLS = [
    ("project_folder",    "Project Folder"),
    ("parcel_num",        "Parcel #"),
    ("control_type",      "Control Type"),
    ("expiration_date",   "Expiration Date"),
    ("acres",             "Acres"),
    ("total_cost_of_land","Total Cost of Land ($)"),
    ("seller_name",       "Seller Name"),
    ("address",           "Address"),
    ("city",              "City"),
    ("state",             "State"),
    ("zip",               "Zip"),
    ("common_ownership",  "Common Ownership"),
    ("annual_lease_debt", "Annual Lease Debt ($)"),
]


def build_excel(summary_rows: list, parcel_rows: list, output_path: str) -> None:
    wb = Workbook()

    # ---- Sheet 1: Site_Summary ----
    ws1 = wb.active
    ws1.title = "Site_Summary"
    ws1.freeze_panes = "A2"

    keys1   = [k for k, _ in SUMMARY_COLS]
    labels1 = [l for _, l in SUMMARY_COLS]
    status_col = keys1.index("status") + 1

    _hrow(ws1, 1, labels1)
    for i, row in enumerate(summary_rows):
        _drow(ws1, i + 2, [row.get(k, "") for k in keys1],
              alt=(i % 2 == 1), status_col=status_col)
    _auto_width(ws1)

    # ---- Sheet 2: Site_Parcels ----
    ws2 = wb.create_sheet("Site_Parcels")
    ws2.freeze_panes = "A2"

    keys2   = [k for k, _ in PARCEL_COLS]
    labels2 = [l for _, l in PARCEL_COLS]

    _hrow(ws2, 1, labels2)
    for i, row in enumerate(parcel_rows):
        _drow(ws2, i + 2, [row.get(k, "") for k in keys2], alt=(i % 2 == 1))
    _auto_width(ws2)

    wb.save(output_path)
    print(f"Excel written to: {output_path}")


# ---------------------------------------------------------------------------
# CLI + MAIN
# ---------------------------------------------------------------------------

def _parse_args():
    p = argparse.ArgumentParser(
        description="Extract Page 3 site info from SC TEB application PDFs."
    )
    p.add_argument("--root",       default=DEFAULT_ROOT,    help="Root folder of extracted project subfolders.")
    p.add_argument("--output-dir", default=DEFAULT_OUT_DIR, help="Output directory for Excel workbook.")
    p.add_argument("--log-dir",    default=DEFAULT_LOG_DIR, help="Directory for run logs.")
    return p.parse_args()


def _parse_project_folder(name: str) -> tuple[str, str]:
    m = re.match(r"^\s*(\d{5})\s*[-–]\s*(.+?)\s*$", name)
    return (m.group(1), m.group(2)) if m else ("", name.strip())


def main():
    args = _parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    os.makedirs(args.log_dir,    exist_ok=True)

    run_ts      = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_xlsx = os.path.join(args.output_dir, f"SC_TEB_P03_SiteInfo_{run_ts}.xlsx")
    latest_xlsx = os.path.join(args.output_dir, "SC_TEB_P03_SiteInfo_latest.xlsx")
    log_path    = os.path.join(args.log_dir,    f"p03_run_{run_ts}.log")

    log_file        = open(log_path, "w", encoding="utf-8")
    original_stdout = sys.stdout

    class Tee:
        def __init__(self, *s): self.streams = s
        def write(self, d):
            for s in self.streams: s.write(d); s.flush()
        def flush(self):
            for s in self.streams: s.flush()

    sys.stdout = Tee(original_stdout, log_file)

    try:
        print(f"Root:       {args.root}")
        print(f"Output dir: {args.output_dir}")
        print()

        if not os.path.isdir(args.root):
            raise FileNotFoundError(f'Root folder not found: "{args.root}"')

        summary_rows = []
        parcel_rows  = []

        for project in sorted(os.listdir(args.root)):
            project_path = os.path.join(args.root, project)
            if not os.path.isdir(project_path):
                continue

            pid, pname = _parse_project_folder(project)

            pdf_path = pick_best_pdf(project_path)
            if not pdf_path:
                print(f"[SKIP] {project} – no PDF found")
                summary_rows.append({"project_folder": project, "status": "No PDF found"})
                continue

            result = parse_page3(pdf_path, project)

            if result["status"] == "OK":
                summary_rows.append(result["summary"])
                for pr in result["parcels"]:
                    parcel_rows.append(pr)
                print(f"[OK]   {project} – "
                      f"lat={result['summary'].get('latitude')} "
                      f"lon={result['summary'].get('longitude')} "
                      f"parcels={len(result['parcels'])}")
            else:
                summary_rows.append({"project_folder": project,
                                     "pdf_path": pdf_path,
                                     "status": result["status"]})
                print(f"[MISS] {project} – {result['status']}")

        build_excel(summary_rows, parcel_rows, output_xlsx)

        import shutil
        shutil.copy2(output_xlsx, latest_xlsx)

        print(f"\nProjects processed: {len(summary_rows)}")
        print(f"Parcels extracted:  {len(parcel_rows)}")
        print(f"\nExcel:  {output_xlsx}")
        print(f"Latest: {latest_xlsx}")
        print(f"Log:    {log_path}")

    finally:
        sys.stdout = original_stdout
        log_file.close()


if __name__ == "__main__":
    main()
