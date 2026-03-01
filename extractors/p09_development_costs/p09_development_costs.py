r"""
SC TEB 2025 — Development Cost Schedule Extractor
==================================================
Walks all project folders under EXTRACTED_ROOT, finds the TEB application PDF in
TAB 1, locates the Development Costs page, extracts the table via Camelot lattice,
maps rows to the canonical 80-line SC Housing form, and writes a single Excel
workbook with:
  - One sheet per project  (named by project ID + short name)
  - A "Status" summary sheet

Columns per project sheet match the Palomino gold standard:
  Line | Item | Development Costs | 4% Basis (Acquisition) |
  4%/9% Basis (New/Rehab) | Eligible Basis (Acq + New/Rehab) |
  Summary of Const Cost Addm | Difference | Raw

Run (Windows PowerShell):
  py -3.14 extract_dev_costs_to_excel.py

Dependencies: camelot-py[cv], pdfplumber, openpyxl
"""

from __future__ import annotations

import atexit
import os
import re
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import camelot
import pdfplumber
import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG  ← edit these paths if needed
# ─────────────────────────────────────────────
EXTRACTED_ROOT = Path(r"C:\Users\tanne\Downloads\SC_TEB\2025\extracted")
OUTPUT_DIR = Path(
    r"C:\Users\tanne\Projects\south-carolina-housing-tools\extractors\p09_development_costs\output"
)
LOG_DIR = Path(
    r"C:\Users\tanne\Projects\south-carolina-housing-tools\extractors\p09_development_costs\logs"
)

# Set to a small integer (e.g. 3) while testing, None for full run
MAX_DEALS: Optional[int] = None

# ─────────────────────────────────────────────
# CAMELOT TEMP DIR SETUP + CLEANUP
# ─────────────────────────────────────────────
_CAMELOT_TEMP = OUTPUT_DIR / "camelot_tmp"
_CAMELOT_TEMP.mkdir(parents=True, exist_ok=True)
for _var in ("TMPDIR", "TEMP", "TMP"):
    os.environ[_var] = str(_CAMELOT_TEMP)


def _safe_rmtree(path: Path, retries: int = 6, sleep_s: float = 0.15) -> None:
    import shutil

    for _ in range(retries):
        try:
            shutil.rmtree(path, ignore_errors=False)
            return
        except PermissionError:
            time.sleep(sleep_s)
        except FileNotFoundError:
            return
    shutil.rmtree(path, ignore_errors=True)


def _install_cleanup() -> None:
    tmp = Path(tempfile.gettempdir())

    def _go():
        for p in tmp.glob("tmp*"):
            if p.is_dir() and any(p.glob("page-*.pdf")):
                _safe_rmtree(p)

    atexit.register(_go)


_install_cleanup()

# ─────────────────────────────────────────────
# CANONICAL 80-LINE FORM
# Keys are lowercase, stripped, with dashes/spaces normalised.
# Values: (line_number, display_label)
# ─────────────────────────────────────────────
CANONICAL: List[Tuple[int, str]] = [
    (1, "Land"),
    (2, "Existing Structures"),
    (3, "Other (Specify)"),  # acquisition other
    (4, "On-Site Improvements"),
    (5, "Off-Site Improvements"),
    (6, "Demolition"),
    (7, "Improvements"),
    (8, "New Construction"),
    (9, "Rehabilitation"),
    (10, "Accessory Structures"),
    (11, "Other Hard Construction Costs"),
    (12, "Furniture, Fixtures, & Equipment"),
    (13, "Contractor Contingency"),
    (14, "General Requirements"),
    (15, "Contractor Profit"),
    (16, "Contractor Overhead"),
    (17, "Architect Fee Design"),
    (18, "Architect Fee Construction Supervision"),
    (19, "Engineering Fees"),
    (20, "Survey"),
    (21, "Real Estate Attorney Fees"),
    (22, "Tax Attorney Fees"),
    (23, "Accountant"),
    (24, "Green Certification"),
    (25, "Other (Specify) - Geotech"),  # soft / other professional
    (26, "Construction Loan Origination Fee"),
    (27, "Construction Loan Interest Paid"),
    (28, "Construction Loan Legal Fees"),
    (29, "Construction Loan Credit Report"),
    (30, "Construction Loan Title & Recording Costs"),
    (31, "Inspection Fees"),
    (32, "Other (Specify) - Bridge Loan Interest"),
    (33, "Construction Insurance"),
    (34, "Performance Bond Premium"),
    (35, "Construction Period Taxes"),
    (36, "Tap Fees and Impact Fees"),
    (37, "Permitting Fees"),
    (38, "Other (Specify)"),  # construction interim other
    (39, "Permanent Loan Origination Fee"),
    (40, "Bond Premium"),
    (41, "Credit Enhancement"),
    (42, "Permanent Loan Title & Recording"),
    (43, "Counsels Fee"),
    (44, "Lenders Counsel Fee"),
    (45, "Appraisal Fees"),
    (46, "Credit Report"),
    (47, "Mortgage Broker Fees"),
    (48, "Permanent Loan Closing"),
    (49, "Underwriter Discount"),
    (50, "Attorney / Legal Fees"),
    (51, "Other (Specify) - Letter of Credit Costs"),
    (52, "Feasibility Study"),
    (53, "Environmental Study"),
    (54, "Market Study"),
    (55, "SC Housing Application Fee"),
    (56, "SC Housing Market Study"),
    (57, "SC Housing Plan/Spec/Site Review"),
    (58, "SC Housing Tax Credit Reservation (10%)"),
    (59, "SC Housing Bond Issuance (0.75%)"),
    (60, "Compliance Fees"),
    (61, "Cost Certification"),
    (62, "Tenant Relocation Costs"),
    (63, "Soil Testing"),
    (64, "Physical Needs Assessment"),
    (65, "Rent-Up Expenses"),
    (66, "Marketing"),  # sometimes labelled separately
    (67, "Other (Specify)"),  # soft costs other
    (68, "Organizational Expenses"),
    (69, "Tax Opinion"),
    (70, "Bridge Loan Fees"),
    (71, "Syndication Fees"),
    (72, "Other (Specify)"),  # syndication other
    (73, "Developer Overhead"),
    (74, "Developer Fee"),
    (75, "Project Consultant Fee"),
    (76, "Other (Specify) - DDF Interest"),
    (77, "Operating Reserves"),
    (78, "Other (Specify)"),  # reserves other
    (79, "COLUMN TOTALS"),
    (80, "TOTAL DEVELOPMENT COST"),
    (81, "TOTAL ELIGIBLE BASIS"),
    (82, "TOTAL INELIGIBLE COSTS"),
]


# Build a quick-lookup by normalised label
def _normalise(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[\s\-_/\\]+", " ", s)
    s = re.sub(r"\s*-\s*-\s*", " ", s)  # collapse trailing "- -" markers
    s = s.strip()
    return s


_CANON_INDEX: Dict[str, int] = {}  # normalised label → line number
for _ln, _label in CANONICAL:
    _CANON_INDEX[_normalise(_label)] = _ln

# ─────────────────────────────────────────────
# SECTION DETECTION
# Groups of lines that map to the same section heading
# ─────────────────────────────────────────────
SECTION_RANGES = [
    (range(1, 4), "Acquisition"),
    (range(4, 8), "Site Work"),
    (range(8, 17), "Rehabilitation and New Construction"),
    (range(17, 26), "Professional Fees"),
    (range(26, 33), "Construction Financing"),
    (range(33, 39), "Construction Interim Costs"),
    (range(39, 52), "Permanent Financing"),
    (range(52, 68), "Soft Costs"),
    (range(68, 73), "Syndication Costs"),
    (range(73, 79), "Developer Fees"),
    (range(77, 79), "Project Reserves"),
    (range(79, 83), "Totals"),
]


def line_to_section(line_num: int) -> str:
    for rng, sec in SECTION_RANGES:
        if line_num in rng:
            return sec
    return ""


# ─────────────────────────────────────────────
# PDF DISCOVERY  (ported from extract_coords.py)
# ─────────────────────────────────────────────
def find_pdfs_recursive(start_dir: Path) -> List[Path]:
    return [p for p in start_dir.rglob("*.pdf") if p.is_file()]


def find_tab1_dir(project_root: Path) -> Optional[Path]:
    candidates: List[Path] = []
    for p in project_root.rglob("*"):
        if not p.is_dir():
            continue
        name = p.name.lower()
        if (
            "tab" in name
            and re.search(r"\b1\b", name)
            and ("app" in name or "application" in name)
        ):
            candidates.append(p)
            continue
        if name.startswith("tab 1") or name.startswith("tab1") or "tab 01" in name:
            candidates.append(p)
            continue
        if re.search(r"^\s*1\s*[-_ ]", name) and (
            "app" in name or "application" in name
        ):
            candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda x: len(x.parts))
    return candidates[0]


def score_pdf_for_application(pdf_path: Path) -> int:
    score = 0
    name = pdf_path.name.lower().strip()
    bad_words = [
        "waiver",
        "map",
        "railroad",
        "site",
        "zoning",
        "utility",
        "environmental",
        "plans",
        "specifications",
        "appraisal",
        "market study",
        "syndication",
        "architect",
        "engineer",
        "certification",
        "site control",
        "checklist",
        "entity",
        "agreement",
        "opinion",
        "certifications",
        "survey",
        "title",
        "easement",
        "phase i",
        "phase 1",
        "soil",
        "geotech",
    ]
    for w in bad_words:
        if w in name:
            score -= 12
    if "executed" in name:
        score -= 25
    if "application page" in name:
        score -= 15
    if "application" in name:
        score += 12
    elif "app" in name:
        score += 6
    if "teb" in name:
        score += 10
    if "teb" in name and ("app" in name or "application" in name):
        score += 20
    if name in ("2025 teb application.pdf", "2026 teb application.pdf"):
        score += 30
    try:
        size_mb = pdf_path.stat().st_size / (1024 * 1024)
        score += min(int(size_mb), 20)
        if size_mb < 1.0:
            score -= 10
    except Exception:
        pass
    return score


def pick_top_pdfs(
    project_root: Path, tab1_dir: Optional[Path], top_n: int = 8
) -> List[Path]:
    pdfs = find_pdfs_recursive(tab1_dir) if tab1_dir else []
    if not pdfs:
        pdfs = find_pdfs_recursive(project_root)
    if not pdfs:
        return []
    scored = [(score_pdf_for_application(p), p) for p in pdfs]
    scored.sort(
        key=lambda x: (x[0], x[1].stat().st_size if x[1].exists() else 0),
        reverse=True,
    )
    return [p for _, p in scored[:top_n]]


# ─────────────────────────────────────────────
# PAGE DETECTION
# ─────────────────────────────────────────────
def _page_lines(pdf_path: Path, max_pages: int = 140) -> List[Tuple[int, List[str]]]:
    pages = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i in range(min(len(pdf.pages), max_pages)):
            txt = (pdf.pages[i].extract_text() or "").replace("\u00a0", " ")
            lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
            pages.append((i, lines))
    return pages


def _score_dev_costs_page(lines: List[str]) -> int:
    text = " ".join(lines).lower()
    score = 0
    if re.search(r"\bdevelopment\s+costs\b", text):
        score += 3
    if re.search(r"\b4%\b", text):
        score += 1
    if re.search(r"\bbas(?:is|i?s|s{2,})\b", text):
        score += 2
    if re.search(r"\bacquisition\b", text):
        score += 2
    if re.search(r"\bnew\s*/\s*rehab\b|\bnew\b.*\brehab\b", text):
        score += 1
    if re.search(r"\baddm\b|\baddendum\b", text):
        score += 1
    if re.search(r"\bdifference\b", text):
        score += 1
    if re.search(r"\bsummary\b.*\bconst\b|\bsummary of const\b", text):
        score += 1
    return score


def find_dev_cost_page(pages: List[Tuple[int, List[str]]]) -> Tuple[int, int]:
    best_page, best_score = -1, 0
    for page_index, lines in pages:
        s = _score_dev_costs_page(lines)
        if s > best_score:
            best_score, best_page = s, page_index
    return (best_page, best_score) if best_score >= 6 else (-1, best_score)


# ─────────────────────────────────────────────
# CAMELOT EXTRACTION
# ─────────────────────────────────────────────
_TABLE_KEYWORDS = [
    "land",
    "site work",
    "on-site",
    "off-site",
    "new construction",
    "rehabilitation",
    "architect",
    "engineering",
    "developer fee",
    "contingency",
    "total development cost",
    "eligible basis",
    "acquisition",
]
MIN_ROWS, MIN_COLS, MAX_COLS = 50, 3, 6


def _score_table(df) -> int:
    text = " ".join(df.astype(str).values.flatten()).lower()
    score = sum(3 for k in _TABLE_KEYWORDS if k in text)
    score += min(len(re.findall(r"\d[\d,]*\.\d{2}|\d[\d,]*", text)) // 50, 6)
    return score


def _extract_best_table(pdf_path: Path, page_0based: int):
    tables = camelot.read_pdf(
        str(pdf_path),
        pages=str(page_0based + 1),
        flavor="lattice",
        line_scale=40,
        copy_text=["v", "h"],
        split_text=True,
    )
    if tables.n == 0:
        return None, 0
    scored = sorted(
        [(s := _score_table(t.df), t.df) for t in tables],
        key=lambda x: x[0],
        reverse=True,
    )
    return scored[0][1], scored[0][0]


def validate_and_extract(pdf_path: Path):
    """
    Returns (ok, reason, dev_page, df) where df is a raw Camelot DataFrame.
    """
    pages = _page_lines(pdf_path)
    dev_page, page_score = find_dev_cost_page(pages)
    if dev_page == -1:
        return False, f"NO_DEV_PAGE (best score={page_score})", None, None

    df, table_score = _extract_best_table(pdf_path, dev_page)
    if df is None:
        return False, "NO_TABLE", dev_page, None

    r, c = df.shape
    if c < MIN_COLS or c > MAX_COLS:
        return False, f"BAD_COLS_{c}", dev_page, None
    if r < MIN_ROWS:
        return False, f"TOO_FEW_ROWS_{r}", dev_page, None

    return True, "OK", dev_page, df


# ─────────────────────────────────────────────
# ROW PARSING
# ─────────────────────────────────────────────
def _clean(cell) -> str:
    return str(cell or "").replace("\u00a0", " ").strip()


def _parse_num(cell: str) -> Optional[float]:
    s = _clean(cell).replace("$", "").replace(",", "").strip()
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1].strip()
    s2 = re.sub(r"[^\d.]", "", s)
    if not s2 or s2.count(".") > 1:
        return None
    try:
        v = float(s2)
        return -v if neg else v
    except Exception:
        return None


def _is_blankish(cells: List[str]) -> bool:
    return all(not c.strip() or re.fullmatch(r"[-–—\s]+", c) for c in cells)


def _is_header(cells: List[str]) -> bool:
    j = " ".join(cells).lower()
    return any(
        h in j
        for h in [
            "development costs",
            "4% (30%)",
            "4% basis",
            "new / rehab",
            "summary of const",
            "cost addm",
            "difference",
        ]
    )


def _split_label_cell(cell: str) -> List[str]:
    """
    Extract meaningful label lines from a multiline cell.
    Drops numeric-only lines, dash-only lines, and empty lines.
    """
    out = []
    for ln in cell.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if re.fullmatch(r"[\d,]+(\.\d+)?", ln):
            continue
        if re.fullmatch(r"[-–—\s]+", ln):
            continue
        if not re.search(r"[A-Za-z]", ln):
            continue
        out.append(ln)
    return out


def _is_section_header_row(cells: List[str]) -> bool:
    """
    Section header rows have the same text repeated across all columns
    and contain no numeric data.
    """
    if len(cells) < 2:
        return False
    c0 = cells[0].strip()
    if not c0 or not re.search(r"[A-Za-z]", c0):
        return False
    # All cols identical → section header
    if all(c.strip() == c0 for c in cells[1:] if c.strip()):
        # No numbers anywhere
        joined = " ".join(cells)
        if not re.search(r"\d[\d,]*", joined):
            return True
    return False


def parse_raw_rows(df) -> List[Dict[str, Any]]:
    """
    Parse the Camelot DataFrame produced by this PDF layout:

    Layout (4 cols):
      col0 = multiline label cell containing ALL item labels for the section,
             repeated identically on every data row within that section.
      col1 = Development Cost value for THIS row's item (one per row)
      col2 = 4% Acquisition basis
      col3 = 4%/9% New/Rehab basis

    Strategy:
      - When we enter a new section group (col0 changes), parse the label list
        from col0 and reset a position counter.
      - Each subsequent row increments the counter → label = label_list[counter].
      - Subtotal rows (col0 empty or numeric-only) are kept as subtotals.
    """
    rows = []
    current_labels: List[str] = []
    current_label_pos: int = 0
    prev_col0: str = ""

    for _, r in df.iterrows():
        cells = [_clean(x) for x in r.tolist()]

        # Skip true header rows (col headers at top of table)
        if _is_header(cells):
            continue

        # Section header rows (e.g. "Acquisition", "Site Work") — reset state
        if _is_section_header_row(cells):
            current_labels = []
            current_label_pos = 0
            prev_col0 = ""
            continue

        col0 = cells[0] if cells else ""
        nums_in_col0 = bool(re.search(r"\d[\d,]{2,}", col0))

        # Subtotal / blank-label rows (col0 empty or pure numeric)
        if not col0.strip() or (not re.search(r"[A-Za-z]", col0) and nums_in_col0):
            # These are section subtotal rows — record with empty label
            dev_cost = _parse_num(cells[1]) if len(cells) > 1 else None
            acq_basis = _parse_num(cells[2]) if len(cells) > 2 else None
            new_rehab = _parse_num(cells[3]) if len(cells) > 3 else None
            if any(v is not None for v in (dev_cost, acq_basis, new_rehab)):
                rows.append(
                    {
                        "label": "_subtotal_",
                        "dev_cost": dev_cost,
                        "acq_basis": acq_basis,
                        "new_rehab": new_rehab,
                        "addm": None,
                        "diff": None,
                        "raw": " | ".join(cells),
                    }
                )
            continue

        # New section group: col0 has changed (new multiline label block)
        col0_key = col0[:80]  # truncate for comparison
        if col0_key != prev_col0:
            current_labels = _split_label_cell(col0)
            current_label_pos = 0
            prev_col0 = col0_key

        # Assign label by position within the group
        if current_label_pos < len(current_labels):
            label = current_labels[current_label_pos]
        else:
            # More rows than labels — use last label + index suffix
            label = (
                current_labels[-1] if current_labels else col0.splitlines()[0].strip()
            )

        current_label_pos += 1

        # Numeric values: col1=dev_cost, col2=acq_basis, col3=new_rehab
        dev_cost = _parse_num(cells[1]) if len(cells) > 1 else None
        acq_basis = _parse_num(cells[2]) if len(cells) > 2 else None
        new_rehab = _parse_num(cells[3]) if len(cells) > 3 else None

        rows.append(
            {
                "label": label,
                "dev_cost": dev_cost,
                "acq_basis": acq_basis,
                "new_rehab": new_rehab,
                "addm": None,  # not present in this 4-col layout
                "diff": None,  # not present in this 4-col layout
                "raw": " | ".join(cells),
            }
        )

    return rows


# ─────────────────────────────────────────────
# CANONICAL LINE MAPPING
# ─────────────────────────────────────────────
def map_to_canonical(raw_rows: List[Dict]) -> List[Dict]:
    """
    Map parsed rows → canonical 80-line form using an ordered, consume-once strategy.

    Because labels like "Other (Specify)" repeat across multiple sections, a simple
    dict lookup would reuse the same parsed row for every occurrence.  Instead we:
      1. Walk the canonical lines in order.
      2. For each canonical line, search the remaining (unconsumed) parsed rows
         from the current position forward for the best label match.
      3. Once a parsed row is consumed it cannot be matched again.

    This preserves section ordering and prevents duplicate assignments.
    """
    # Filter out subtotal rows for matching purposes
    matchable = [r for r in raw_rows if r.get("label") != "_subtotal_"]

    # Track which parsed rows have been consumed (by index)
    consumed: set = set()

    def _match_score(canon_key: str, parsed_key: str) -> int:
        """Return match quality: 2=exact, 1=substring, 0=no match."""
        if canon_key == parsed_key:
            return 2
        if canon_key in parsed_key or parsed_key in canon_key:
            return 1
        return 0

    def _find_next(canon_key: str, search_from: int) -> Tuple[Optional[Dict], int]:
        """
        Find the best unconsumed match for canon_key starting at search_from.
        Returns (matched_row, index) or (None, -1).
        Prefers exact matches; falls back to substring match.
        """
        # First pass: exact match
        for i in range(search_from, len(matchable)):
            if i in consumed:
                continue
            pk = _normalise(matchable[i]["label"])
            if _match_score(canon_key, pk) == 2:
                return matchable[i], i

        # Second pass: substring match across full remaining list
        for i in range(search_from, len(matchable)):
            if i in consumed:
                continue
            pk = _normalise(matchable[i]["label"])
            if _match_score(canon_key, pk) == 1:
                return matchable[i], i

        return None, -1

    output_rows = []

    for line_num, label in CANONICAL:
        canon_key = _normalise(label)
        matched, idx = _find_next(
            canon_key, 0
        )  # always search from beginning of unconsumed

        if matched is not None:
            consumed.add(idx)

        output_rows.append(
            {
                "line": line_num,
                "label": label,
                "dev_cost": matched["dev_cost"] if matched else None,
                "acq_basis": matched["acq_basis"] if matched else None,
                "new_rehab": matched["new_rehab"] if matched else None,
                "addm": matched["addm"] if matched else None,
                "diff": matched["diff"] if matched else None,
                "raw": matched["raw"] if matched else "",
            }
        )

    return output_rows


# ─────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────
COLS = [
    "Line",
    "Item",
    "Development Costs",
    "4% Basis (Acquisition)",
    "4%/9% Basis (New/Rehab)",
    "Eligible Basis (Acq + New/Rehab)",
    "Summary of Const Cost Addm",
    "Difference",
    "Raw",
]

# Colour palette
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")  # dark blue
_SECTION_FILL = PatternFill("solid", start_color="D6E4F0")  # light blue
_TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")  # light yellow
_ALT_FILL = PatternFill("solid", start_color="F5F5F5")  # very light grey
_BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
_SECTION_FONT = Font(name="Arial", bold=True, size=9)
_DATA_FONT = Font(name="Arial", size=9)
_TOTAL_FONT = Font(name="Arial", bold=True, size=9)
_NUM_FMT = '#,##0;(#,##0);"-"'

SECTION_HEADER_ROWS = {
    1: "Acquisition",
    4: "Site Work",
    8: "Rehabilitation and New Construction",
    17: "Professional Fees",
    26: "Construction Financing",
    33: "Construction Interim Costs",
    39: "Permanent Financing",
    52: "Soft Costs",
    68: "Syndication Costs",
    73: "Developer Fees",
    77: "Project Reserves",
    79: "── Totals ──",
}
TOTAL_LINES = {79, 80, 81, 82}
NUM_COLS = {3, 4, 5, 6, 7, 8}  # 1-indexed Excel col positions for numeric data


def _write_project_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    project_id: str,
    project_name: str,
    mapped_rows: List[Dict],
    raw_rows: List[Dict],
) -> None:
    ws = wb.create_sheet(title=sheet_name)

    # ── Title row ──
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{project_id} – {project_name} | Development Cost Schedule"
    title_cell.font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # ── Column headers ──
    for col_idx, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER_THIN
    ws.row_dimensions[2].height = 28

    excel_row = 3
    prev_section = None

    for mapped in mapped_rows:
        line = mapped["line"]
        label = mapped["label"]

        # ── Section header row ──
        section_label = SECTION_HEADER_ROWS.get(line)
        if section_label and section_label != prev_section:
            ws.merge_cells(
                start_row=excel_row,
                start_column=1,
                end_row=excel_row,
                end_column=len(COLS),
            )
            sec_cell = ws.cell(row=excel_row, column=1, value=section_label)
            sec_cell.font = _SECTION_FONT
            sec_cell.fill = _SECTION_FILL
            sec_cell.alignment = Alignment(
                horizontal="left", vertical="center", indent=1
            )
            sec_cell.border = _BORDER_THIN
            ws.row_dimensions[excel_row].height = 16
            excel_row += 1
            prev_section = section_label

        # ── Data row ──
        is_total = line in TOTAL_LINES
        fill = _TOTAL_FILL if is_total else (_ALT_FILL if excel_row % 2 == 0 else None)
        font = _TOTAL_FONT if is_total else _DATA_FONT

        row_data = [
            line if line <= 78 else "",  # hide totals line numbers
            label,
            mapped["dev_cost"],
            mapped["acq_basis"],
            mapped["new_rehab"],
            # Eligible Basis = Acq + New/Rehab (formula if both present, else whichever)
            None,  # placeholder — set below
            mapped["addm"],
            mapped["diff"],
            mapped["raw"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = font
            if fill:
                cell.fill = fill
            cell.border = _BORDER_THIN
            cell.alignment = Alignment(
                vertical="center", wrap_text=(col_idx == len(COLS))
            )

            if col_idx == 6:
                # Eligible Basis: sum of acq + new/rehab columns (D + E)
                d_col = get_column_letter(4)
                e_col = get_column_letter(5)
                cell.value = f'=IFERROR({d_col}{excel_row}+{e_col}{excel_row},"")'
                cell.number_format = _NUM_FMT
            elif col_idx in NUM_COLS and col_idx != 6:
                cell.value = value
                cell.number_format = _NUM_FMT
            else:
                cell.value = value

        ws.row_dimensions[excel_row].height = 14
        excel_row += 1

    # ── Column widths ──
    ws.column_dimensions["A"].width = 6  # Line
    ws.column_dimensions["B"].width = 42  # Item
    ws.column_dimensions["C"].width = 18  # Dev Costs
    ws.column_dimensions["D"].width = 18  # Acq Basis
    ws.column_dimensions["E"].width = 18  # New/Rehab
    ws.column_dimensions["F"].width = 20  # Eligible Basis
    ws.column_dimensions["G"].width = 18  # Addm
    ws.column_dimensions["H"].width = 14  # Difference
    ws.column_dimensions["I"].width = 50  # Raw

    ws.freeze_panes = "C3"


def _write_status_sheet(wb: openpyxl.Workbook, status_rows: List[Dict]) -> None:
    ws = wb.create_sheet(title="Status", index=0)
    headers = [
        "project_folder",
        "project_id",
        "project_name",
        "status",
        "pdf_path",
        "dev_costs_page",
        "notes",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER_THIN
        cell.alignment = Alignment(horizontal="center")

    ok_fill = PatternFill("solid", start_color="C6EFCE")
    err_fill = PatternFill("solid", start_color="FFC7CE")

    for r_idx, row in enumerate(status_rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=row.get(key, ""))
            cell.font = _DATA_FONT
            cell.border = _BORDER_THIN
            if key == "status":
                cell.fill = ok_fill if row.get("status") == "OK" else err_fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 60


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def _parse_project_folder(name: str) -> Tuple[str, str]:
    m = re.match(r"^\s*(\d{5})\s*[-–]\s*(.+?)\s*$", name)
    return (m.group(1), m.group(2)) if m else ("", name.strip())


def _safe_sheet_name(project_id: str, project_name: str) -> str:
    raw = f"{project_id} {project_name}"
    safe = re.sub(r"[\\/*?:\[\]]", "", raw)[:31]
    return safe


def run():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    out_xlsx = OUTPUT_DIR / f"SC_TEB_DevCosts_{ts}.xlsx"
    latest_xlsx = OUTPUT_DIR / "SC_TEB_DevCosts_latest.xlsx"
    log_path = LOG_DIR / f"run_dev_costs_{ts}.log"
    _log_file = open(log_path, "w", encoding="utf-8")

    import sys as _sys
    _orig_stdout = _sys.stdout

    class _Tee:
        def __init__(self, *s): self.streams = s
        def write(self, d):
            for s in self.streams: s.write(d); s.flush()
        def flush(self):
            for s in self.streams: s.flush()

    _sys.stdout = _Tee(_orig_stdout, _log_file)

    projects = sorted(p for p in EXTRACTED_ROOT.iterdir() if p.is_dir())
    if MAX_DEALS is not None:
        projects = projects[:MAX_DEALS]
    print(f"Found {len(projects)} project folders under: {EXTRACTED_ROOT}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    status_rows: List[Dict] = []

    for project_path in projects:
        folder = project_path.name
        pid, pname = _parse_project_folder(folder)
        meta = {"project_folder": folder, "project_id": pid, "project_name": pname}

        tab1 = find_tab1_dir(project_path)
        candidates = pick_top_pdfs(project_path, tab1)

        if not candidates:
            print(f"[SKIP] {folder} — no PDFs found")
            status_rows.append(
                {
                    **meta,
                    "status": "MISS_NO_PDF",
                    "pdf_path": "",
                    "dev_costs_page": "",
                    "notes": "",
                }
            )
            continue

        chosen_pdf = chosen_page = chosen_df = None
        attempt_notes = []

        for cand in candidates:
            ok, reason, dev_page, df = validate_and_extract(cand)
            attempt_notes.append(f"{cand.name}=>{reason}")
            if ok:
                chosen_pdf, chosen_page, chosen_df = cand, dev_page, df
                break

        if chosen_df is None:
            note = " | ".join(attempt_notes[:5])
            print(f"[MISS] {folder} — {note}")
            status_rows.append(
                {
                    **meta,
                    "status": "MISS_DEV_COSTS_NOT_VALIDATED",
                    "pdf_path": str(candidates[0]),
                    "dev_costs_page": "",
                    "notes": note,
                }
            )
            continue

        try:
            raw_rows = parse_raw_rows(chosen_df)
            mapped_rows = map_to_canonical(raw_rows)
            sheet_name = _safe_sheet_name(pid, pname)
            _write_project_sheet(wb, sheet_name, pid, pname, mapped_rows, raw_rows)

            note = (
                f"pdf={chosen_pdf.name}; page={chosen_page}; "
                f"raw_rows={len(raw_rows)}; mapped={len(mapped_rows)}"
            )
            print(f"[OK]   {folder} — {note}")
            status_rows.append(
                {
                    **meta,
                    "status": "OK",
                    "pdf_path": str(chosen_pdf),
                    "dev_costs_page": chosen_page,
                    "notes": note,
                }
            )
        except Exception as e:
            print(f"[ERR]  {folder} — {e}")
            status_rows.append(
                {
                    **meta,
                    "status": f"ERROR: {e}",
                    "pdf_path": str(chosen_pdf),
                    "dev_costs_page": chosen_page,
                    "notes": "",
                }
            )

    _write_status_sheet(wb, status_rows)
    wb.save(str(out_xlsx))

    import shutil

    shutil.copyfile(out_xlsx, latest_xlsx)

    ok_count = sum(1 for r in status_rows if r["status"] == "OK")
    print(f"\nDONE — {ok_count}/{len(status_rows)} projects extracted")
    print(f"Output: {out_xlsx}")
    print(f"Latest: {latest_xlsx}")
    print(f"Log:    {log_path}")

    _sys.stdout = _orig_stdout
    _log_file.close()


if __name__ == "__main__":
    run()
